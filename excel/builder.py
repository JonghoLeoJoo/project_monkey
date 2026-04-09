"""
Excel Builder - orchestrates workbook creation by sheet modules.

All dollar amounts displayed in $USD millions (raw values divide by 1,000,000).
"""

from openpyxl import Workbook
from typing import Dict

from .sheet_financial_statements import _write_financial_statements
from .sheet_wacc import _write_wacc_sheet
from .sheet_dcf import _write_dcf_model
from .sheet_validation import _write_validation_sheet, _check_status


def create_excel(company_info: Dict, financial_data: Dict, output_file: str):
    """Create the full financial model workbook and save it."""
    wb = Workbook()

    ws_fs = wb.active
    ws_fs.title = 'Financial Statements'

    ws_wacc = wb.create_sheet('WACC')
    ws_dcf = wb.create_sheet('DCF Model')
    ws_val = wb.create_sheet('Data Validation')

    print("  Building Financial Statements sheet...")
    fs_rows = _write_financial_statements(ws_fs, company_info, financial_data)

    print("  Building WACC sheet...")
    wacc_rows = _write_wacc_sheet(ws_wacc, company_info, financial_data, fs_rows)

    print("  Building DCF Model sheet...")
    _write_dcf_model(ws_dcf, company_info, financial_data, fs_rows, wacc_rows)

    print("  Building Data Validation sheet...")
    validation_results = _write_validation_sheet(
        ws_val, company_info, financial_data, fs_rows)

    wb.save(output_file)
    print(f"  Saved: {output_file}")

    return validation_results


__all__ = [
    'create_excel',
    '_check_status',
]
