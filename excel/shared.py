"""
Shared styles, constants, and utility helpers for Excel builder sheets.
"""

from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from typing import Dict, Optional, List
import statistics

# ── Palette ───────────────────────────────────────────────────────────────────
DARK_BLUE   = "1F4E79"
MED_BLUE    = "2E75B6"
LIGHT_BLUE  = "BDD7EE"
XLIGHT_BLUE = "DEEAF1"
DARK_GREEN  = "375623"
LIGHT_GREEN = "E2EFDA"
YELLOW      = "FFF2CC"
DARK_YELLOW = "F4B942"
LIGHT_RED   = "FFE2CC"
WHITE       = "FFFFFF"
LIGHT_GRAY  = "F2F2F2"
MED_GRAY    = "D6D6D6"

# ── Borders ───────────────────────────────────────────────────────────────────
_thin   = Side(style='thin',   color='000000')
_medium = Side(style='medium', color='000000')
_thick  = Side(style='thick',  color='000000')
THIN_BOX  = Border(left=_thin,   right=_thin,   top=_thin,   bottom=_thin)
BOT_MED   = Border(bottom=_medium)
BOT_THICK = Border(bottom=_thick)
TOP_THIN  = Border(top=_thin)

# ── Number formats ────────────────────────────────────────────────────────────
FMT_DOLLAR  = '#,##0.0'        # e.g. 1,234.5  ($M)
FMT_DOLLAR2 = '#,##0.00'       # EPS
FMT_PCT     = '0.0%'
FMT_MULT    = '0.0x'
FMT_INT     = '#,##0'

__all__ = [
    # Palette
    'DARK_BLUE', 'MED_BLUE', 'LIGHT_BLUE', 'XLIGHT_BLUE', 'DARK_GREEN',
    'LIGHT_GREEN', 'YELLOW', 'DARK_YELLOW', 'LIGHT_RED', 'WHITE',
    'LIGHT_GRAY', 'MED_GRAY',
    # Borders
    'THIN_BOX', 'BOT_MED', 'BOT_THICK', 'TOP_THIN',
    # Number formats
    'FMT_DOLLAR', 'FMT_DOLLAR2', 'FMT_PCT', 'FMT_MULT', 'FMT_INT',
    # Style helpers
    '_fill', '_font', '_align', '_style', '_set_col_widths',
    # Data helpers
    '_val', '_safe_avg',
    # Cell-writing helpers
    '_write_section_header', '_write_col_headers', '_write_row', '_spacer',
    # openpyxl re-exports (used directly in sheet modules)
    'Font', 'PatternFill', 'Alignment', 'Border', 'Side',
    'Comment', 'get_column_letter', 'DataValidation',
]


def _fill(hex_color: str) -> PatternFill:
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')


def _font(bold=False, color='000000', size=10, italic=False) -> Font:
    return Font(bold=bold, color=color, size=size, italic=italic, name='Calibri')


def _align(h='left', v='center', wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _style(cell, fill_hex=None, bold=False, font_color='000000',
           h_align='left', number_format=None, border=None, italic=False):
    if fill_hex:
        cell.fill = _fill(fill_hex)
    cell.font  = _font(bold=bold, color=font_color, italic=italic)
    cell.alignment = _align(h=h_align)
    if number_format:
        cell.number_format = number_format
    if border:
        cell.border = border


def _set_col_widths(ws, widths: Dict[int, float]):
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def _val(d: Dict[int, Optional[float]], year: int,
         scale: float = 1e6, negate: bool = False) -> Optional[float]:
    """Return value scaled to millions; return None if missing."""
    v = d.get(year)
    if v is None:
        return None
    v = v / scale
    return -v if negate else v


def _safe_avg(values: List[Optional[float]]) -> float:
    cleaned = [v for v in values if v is not None]
    return statistics.mean(cleaned) if cleaned else 0.0


def _write_section_header(ws, row: int, title: str, cols: int = 6):
    cell = ws.cell(row=row, column=1, value=title)
    _style(cell, fill_hex=MED_BLUE, bold=True, font_color=WHITE, h_align='left')
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell.border = THIN_BOX


def _write_col_headers(ws, row: int, year_cols: List[int], years: List[int],
                       start_col: int = 2):
    ws.cell(row=row, column=1).value = ''
    for i, yr in enumerate(years):
        col = start_col + i
        c = ws.cell(row=row, column=col, value=f'FY{yr}')
        _style(c, fill_hex=DARK_BLUE, bold=True, font_color=WHITE, h_align='center')
        c.border = THIN_BOX


def _write_row(ws, row: int, label: str, data: Dict[int, Optional[float]],
               years: List[int], start_col: int = 2, scale: float = 1e6,
               fmt: str = FMT_DOLLAR, bold: bool = False, fill: str = None,
               negate: bool = False, indent: int = 0) -> int:
    """Write one labelled data row. Returns the row number."""
    prefix = '  ' * indent
    c = ws.cell(row=row, column=1, value=prefix + label)
    _style(c, fill_hex=fill, bold=bold)

    for i, yr in enumerate(sorted(years)):
        col = start_col + i
        v = _val(data, yr, scale=scale, negate=negate)
        cell = ws.cell(row=row, column=col, value=v)
        _style(cell, fill_hex=fill or (LIGHT_GRAY if i % 2 == 0 else WHITE),
               bold=bold, h_align='right', number_format=fmt)
        if bold:
            cell.border = BOT_MED
    return row


def _spacer(ws, row: int, cols: int = 6):
    for c in range(1, cols + 1):
        ws.cell(row=row, column=c).value = None
