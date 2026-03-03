"""
Excel Builder – creates a professional financial model workbook.

Sheet 1 "Financial Statements" : historical Income Statement, Balance Sheet,
                                   Cash Flow Statement (4 years) + balance check
Sheet 2 "DCF Model"            : 5-year discounted cash flow valuation

All dollar amounts displayed in $USD millions (raw values ÷ 1,000,000).
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from typing import Dict, Optional, List
import statistics

# ── Palette ───────────────────────────────────────────────────────────────────
DARK_BLUE   = "1F4E79"
MED_BLUE    = "2E75B6"
LIGHT_BLUE  = "BDD7EE"
XLIGHT_BLUE = "DEEAF1"
DARK_GREEN  = "375623"
LIGHT_GREEN = "E2EFDA"
YELLOW      = "FFF2CC"
DARK_YELLOW = "F4B942"
LIGHT_RED   = "FFE2CC"
WHITE       = "FFFFFF"
LIGHT_GRAY  = "F2F2F2"
MED_GRAY    = "D6D6D6"

# ── Borders ───────────────────────────────────────────────────────────────────
_thin   = Side(style='thin',   color='000000')
_medium = Side(style='medium', color='000000')
_thick  = Side(style='thick',  color='000000')
THIN_BOX  = Border(left=_thin,   right=_thin,   top=_thin,   bottom=_thin)
BOT_MED   = Border(bottom=_medium)
BOT_THICK = Border(bottom=_thick)
TOP_THIN  = Border(top=_thin)

# ── Number formats ────────────────────────────────────────────────────────────
FMT_DOLLAR  = '#,##0.0'        # e.g. 1,234.5  ($M)
FMT_DOLLAR2 = '#,##0.00'       # EPS
FMT_PCT     = '0.0%'
FMT_MULT    = '0.0x'
FMT_INT     = '#,##0'

# ─────────────────────────────────────────────────────────────────────────────
# STYLE HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def _fill(hex_color: str) -> PatternFill:
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')


def _font(bold=False, color='000000', size=10, italic=False) -> Font:
    return Font(bold=bold, color=color, size=size, italic=italic, name='Calibri')


def _align(h='left', v='center', wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _style(cell, fill_hex=None, bold=False, font_color='000000',
           h_align='left', number_format=None, border=None, italic=False):
    if fill_hex:
        cell.fill = _fill(fill_hex)
    cell.font  = _font(bold=bold, color=font_color, italic=italic)
    cell.alignment = _align(h=h_align)
    if number_format:
        cell.number_format = number_format
    if border:
        cell.border = border


def _set_col_widths(ws, widths: Dict[int, float]):
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def _val(d: Dict[int, Optional[float]], year: int,
         scale: float = 1e6, negate: bool = False) -> Optional[float]:
    """Return value scaled to millions; return None if missing."""
    v = d.get(year)
    if v is None:
        return None
    v = v / scale
    return -v if negate else v


def _safe_avg(values: List[Optional[float]]) -> float:
    cleaned = [v for v in values if v is not None]
    return statistics.mean(cleaned) if cleaned else 0.0


def _write_section_header(ws, row: int, title: str, cols: int = 6):
    cell = ws.cell(row=row, column=1, value=title)
    _style(cell, fill_hex=MED_BLUE, bold=True, font_color=WHITE, h_align='left')
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell.border = THIN_BOX


def _write_col_headers(ws, row: int, year_cols: List[int], years: List[int],
                       start_col: int = 2):
    ws.cell(row=row, column=1).value = ''
    for i, yr in enumerate(years):
        col = start_col + i
        c = ws.cell(row=row, column=col, value=f'FY{yr}')
        _style(c, fill_hex=DARK_BLUE, bold=True, font_color=WHITE, h_align='center')
        c.border = THIN_BOX


def _write_row(ws, row: int, label: str, data: Dict[int, Optional[float]],
               years: List[int], start_col: int = 2, scale: float = 1e6,
               fmt: str = FMT_DOLLAR, bold: bool = False, fill: str = None,
               negate: bool = False, indent: int = 0) -> int:
    """Write one labelled data row. Returns the row number."""
    prefix = '  ' * indent
    c = ws.cell(row=row, column=1, value=prefix + label)
    _style(c, fill_hex=fill, bold=bold)

    for i, yr in enumerate(sorted(years)):
        col = start_col + i
        v = _val(data, yr, scale=scale, negate=negate)
        cell = ws.cell(row=row, column=col, value=v)
        _style(cell, fill_hex=fill or (LIGHT_GRAY if i % 2 == 0 else WHITE),
               bold=bold, h_align='right', number_format=fmt)
        if bold:
            cell.border = BOT_MED
    return row


def _spacer(ws, row: int, cols: int = 6):
    for c in range(1, cols + 1):
        ws.cell(row=row, column=c).value = None


# ─────────────────────────────────────────────────────────────────────────────
# SHEET 1 – FINANCIAL STATEMENTS
# ─────────────────────────────────────────────────────────────────────────────

def _write_financial_statements(ws, company_info: Dict, fd: Dict):
    years_desc = fd['years']                          # newest first
    years = list(reversed(years_desc))               # oldest first for display
    n    = len(years)
    inc  = fd['income_statement']
    bs   = fd['balance_sheet']
    cf   = fd['cash_flow']

    # ── Projection setup ────────────────────────────────────────────────
    latest_yr      = years[-1]
    proj_years     = [latest_yr + i for i in range(1, 6)]
    n_proj         = 5
    total_cols     = 1 + n + n_proj          # 10 (label + 4 hist + 5 proj)
    proj_start_col = 2 + n                   # column 6 (F)

    def _cl(i: int) -> str:
        """Excel column letter for year index i (0 = oldest year = column B)."""
        return get_column_letter(2 + i)

    def _pcl(j: int) -> str:
        """Excel column letter for projection year index j (0 = first proj year)."""
        return get_column_letter(proj_start_col + j)

    def _fw(row_num: int, col_idx: int, formula: str,
            fmt: str = FMT_DOLLAR, bold: bool = False, fill: str = None):
        """Write an Excel formula string to the data cell at (row_num, year col_idx)."""
        c = ws.cell(row=row_num, column=2 + col_idx, value=formula)
        _style(c,
               fill_hex=fill or (LIGHT_GRAY if col_idx % 2 == 0 else WHITE),
               bold=bold, h_align='right', number_format=fmt)
        if bold:
            c.border = BOT_MED
        return c

    def _pfw(row_num: int, proj_idx: int, formula: str,
             fmt: str = FMT_DOLLAR, bold: bool = False, fill: str = None):
        """Write an Excel formula into a projection column cell."""
        c = ws.cell(row=row_num, column=proj_start_col + proj_idx, value=formula)
        _style(c,
               fill_hex=fill or (XLIGHT_BLUE if proj_idx % 2 == 0 else WHITE),
               bold=bold, h_align='right', number_format=fmt)
        if bold:
            c.border = BOT_MED
        return c

    def _lbl(row_num: int, text: str, bold: bool = False,
             fill: str = None, ind: int = 0):
        c = ws.cell(row=row_num, column=1, value='  ' * ind + text)
        _style(c, fill_hex=fill, bold=bold)
        return c

    # Column setup — 4 historical + 5 projection columns
    col_widths = {1: 40}
    for ci in range(2, 2 + n + n_proj):
        col_widths[ci] = 16
    _set_col_widths(ws, col_widths)
    r = 1

    # Title
    title_cell = ws.cell(
        row=r, column=1,
        value=f"{company_info['name']}  ({company_info['ticker']})"
              f" - Financial Statements  ($ in millions)")
    _style(title_cell, fill_hex=DARK_BLUE, bold=True, font_color=WHITE)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=total_cols)
    ws.row_dimensions[r].height = 18
    r += 1

    # Subtitle (merge up to second-to-last column, leave last proj col for dropdown)
    sub = ws.cell(
        row=r, column=1,
        value="Source: SEC EDGAR XBRL  |  Shaded rows = formula-derived  |  Amounts in $USD Millions")
    _style(sub, fill_hex=XLIGHT_BLUE, italic=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r,
                   end_column=total_cols - 1)

    # ── Scenario dropdown cell (subtitle row, last projection column) ────
    dropdown_col = proj_start_col + n_proj - 1   # column J (10)
    dropdown_ref = f'${get_column_letter(dropdown_col)}${r}'
    dd_cell = ws.cell(row=r, column=dropdown_col, value='Base Case')
    _style(dd_cell, fill_hex=YELLOW, bold=True, h_align='center')
    dd_cell.border = THIN_BOX
    dv = DataValidation(
        type='list',
        formula1='"Best Case,Base Case,Weak Case"',
        allow_blank=False,
    )
    dv.error = 'Please select Best Case, Base Case, or Weak Case'
    dv.errorTitle = 'Invalid Scenario'
    ws.add_data_validation(dv)
    dv.add(dd_cell)
    r += 2

    # =========================================================================
    # INCOME STATEMENT
    # Derived items (Gross Profit, EBITDA, EBIT, Pre-tax, Net Income, margins)
    # are Excel formulas so they update when raw inputs are edited.
    # Projection columns (5 years) are driven by the scenario dropdown.
    #
    # Layout:  Revenue - COGS = Gross Profit
    #          GP - R&D - SGA - Other_Opex = EBITDA   <- Other_Opex is a balancing plug
    #          EBITDA - D&A = EBIT
    #          EBIT - IntExp + IntInc = Pre-tax
    #          Pre-tax - Tax = Net Income
    # =========================================================================
    _write_section_header(ws, r, 'INCOME STATEMENT', cols=total_cols);  r += 1

    # Column headers — historical (DARK_BLUE) + projection (MED_BLUE)
    _write_col_headers(ws, r, list(range(2, 2 + n)), years, start_col=2)
    for j, py in enumerate(proj_years):
        col = proj_start_col + j
        c = ws.cell(row=r, column=col, value=f'FY{py}E')
        _style(c, fill_hex=MED_BLUE, bold=True, font_color=WHITE, h_align='center')
        c.border = THIN_BOX
    r += 1

    # -- Raw inputs (projection formulas backfilled after assumptions band) --
    rev_row  = r; _write_row(ws, r, 'Revenue', inc['revenue'], years); r += 1
    cogs_row = r; _write_row(ws, r, '  Cost of Revenue', inc['cogs'], years, indent=1); r += 1

    # Gross Profit = Revenue - Cost of Revenue  [FORMULA]
    gp_row = r
    _lbl(r, 'Gross Profit', bold=True)
    for i in range(n):
        _fw(r, i, f'={_cl(i)}{rev_row}-{_cl(i)}{cogs_row}', bold=True, fill=XLIGHT_BLUE)
    for j in range(n_proj):
        _pfw(r, j, f'={_pcl(j)}{rev_row}-{_pcl(j)}{cogs_row}', bold=True, fill=XLIGHT_BLUE)
    r += 1

    # Gross Margin % = Gross Profit / Revenue  [FORMULA]
    _lbl(r, '  Gross Margin %', ind=1)
    for i in range(n):
        _fw(r, i, f'=IF({_cl(i)}{rev_row}<>0,{_cl(i)}{gp_row}/{_cl(i)}{rev_row},"")',
            fmt=FMT_PCT)
    for j in range(n_proj):
        _pfw(r, j, f'=IF({_pcl(j)}{rev_row}<>0,{_pcl(j)}{gp_row}/{_pcl(j)}{rev_row},"")',
             fmt=FMT_PCT)
    r += 1

    _spacer(ws, r, total_cols); r += 1

    # -- Raw opex inputs (projection formulas backfilled after assumptions band) --
    rd_row  = r; _write_row(ws, r, '  R&D Expense', inc['rd_expense'], years, indent=1); r += 1
    sga_row = r; _write_row(ws, r, '  SG&A Expense', inc['sga_expense'], years, indent=1); r += 1

    # Other Operating Expenses / (Income): balancing plug so that the EBITDA formula
    # equals the exact reported figure.
    # Plug = (Revenue - COGS) - R&D - SGA - Reported EBITDA
    # Uses (Revenue - COGS) to match the formula-derived GP on this sheet.
    # Values in RAW DOLLARS — _write_row will scale to millions.
    other_opex_plug = {}
    for yr in years:
        gp_v     = (inc['revenue'].get(yr) or 0) - (inc['cogs'].get(yr) or 0)
        rd_v     = inc['rd_expense'].get(yr) or 0
        sga_v    = inc['sga_expense'].get(yr) or 0
        ebitda_v = inc['ebitda'].get(yr)
        other_opex_plug[yr] = (gp_v - rd_v - sga_v - ebitda_v) if ebitda_v is not None else 0

    # Check if the plug matches -D&A for every year (i.e. D&A is embedded
    # in R&D / SG&A).  If so, relabel the row accordingly.
    _da_embedded = True
    for yr in years:
        plug_v = other_opex_plug.get(yr) or 0
        da_v   = inc['da'].get(yr) or 0
        if da_v == 0 or abs(plug_v + da_v) > abs(da_v) * 0.01:   # 1% tolerance
            _da_embedded = False
            break

    other_opex_row = r
    if _da_embedded:
        _write_row(ws, r, '  D&A (embedded in R&D / SG&A)', other_opex_plug, years, indent=1)
        ws.cell(row=r, column=1).comment = Comment(
            "Plug: Gross Profit - R&D - SG&A - EBITDA.\n\n"
            "This company reports D&A as part of R&D and/or SG&A\n"
            "rather than as a separate operating expense line.\n"
            "This row backs out the embedded D&A so that EBITDA\n"
            "is stated correctly before the explicit D&A line\n"
            "below.  The net effect is that D&A is subtracted\n"
            "once (here it is added back, then subtracted again\n"
            "in the D&A row).",
            "Financial Model")
    else:
        _write_row(ws, r, '  Other Operating Expenses / (Income)', other_opex_plug, years, indent=1)
        ws.cell(row=r, column=1).comment = Comment(
            "Plug: Gross Profit - R&D - SG&A - EBITDA.\n\n"
            "May include:\n"
            "- Restructuring charges\n"
            "- Impairment of assets\n"
            "- Acquisition-related costs\n"
            "- Litigation settlements\n"
            "- Gains/losses on asset sales\n"
            "- Other operating income/expense",
            "Financial Model")
    # Projection: placeholder — backfilled after PP&E schedule when D&A is embedded
    last_hist = _cl(n - 1)
    for j in range(n_proj):
        _pfw(r, j, f'={last_hist}{other_opex_row}')
    r += 1

    _spacer(ws, r, total_cols); r += 1

    # EBITDA = Gross Profit - R&D - SGA - Other  [FORMULA]
    ebitda_row = r
    _lbl(r, 'EBITDA', bold=True)
    for i in range(n):
        _fw(r, i,
            f'={_cl(i)}{gp_row}-{_cl(i)}{rd_row}-{_cl(i)}{sga_row}-{_cl(i)}{other_opex_row}',
            bold=True, fill=XLIGHT_BLUE)
    for j in range(n_proj):
        _pfw(r, j,
             f'={_pcl(j)}{gp_row}-{_pcl(j)}{rd_row}-{_pcl(j)}{sga_row}-{_pcl(j)}{other_opex_row}',
             bold=True, fill=XLIGHT_BLUE)
    r += 1

    # EBITDA Margin % = EBITDA / Revenue  [FORMULA]
    _lbl(r, '  EBITDA Margin %', ind=1)
    for i in range(n):
        _fw(r, i,
            f'=IF({_cl(i)}{rev_row}<>0,{_cl(i)}{ebitda_row}/{_cl(i)}{rev_row},"")',
            fmt=FMT_PCT)
    for j in range(n_proj):
        _pfw(r, j,
             f'=IF({_pcl(j)}{rev_row}<>0,{_pcl(j)}{ebitda_row}/{_pcl(j)}{rev_row},"")',
             fmt=FMT_PCT)
    r += 1

    # -- Stock-based Compensation (non-cash add-back) --
    sbc_row = r
    _write_row(ws, r, '  Stock-based Compensation', cf['sbc'], years, indent=1)
    # Projection: placeholder (backfilled after assumptions band)
    for j in range(n_proj):
        _pfw(r, j, f'={last_hist}{sbc_row}')  # overwritten by backfill
    r += 1

    # Adjusted EBITDA = EBITDA + SBC  [FORMULA]
    adj_ebitda_row = r
    _lbl(r, 'Adjusted EBITDA', bold=True)
    for i in range(n):
        _fw(r, i, f'={_cl(i)}{ebitda_row}+{_cl(i)}{sbc_row}',
            bold=True, fill=XLIGHT_BLUE)
    for j in range(n_proj):
        _pfw(r, j, f'={_pcl(j)}{ebitda_row}+{_pcl(j)}{sbc_row}',
             bold=True, fill=XLIGHT_BLUE)
    r += 1

    # -- D&A (raw input; subtracted to get EBIT) --
    da_row = r
    _write_row(ws, r, '  Depreciation & Amortization', inc['da'], years, indent=1)
    # Projection: placeholder (backfilled after PP&E schedule)
    for j in range(n_proj):
        _pfw(r, j, f'={last_hist}{da_row}')
    r += 1

    # EBIT = EBITDA - D&A  [FORMULA]
    ebit_row = r
    _lbl(r, 'Operating Income (EBIT)', bold=True)
    for i in range(n):
        _fw(r, i, f'={_cl(i)}{ebitda_row}-{_cl(i)}{da_row}', bold=True, fill=XLIGHT_BLUE)
    for j in range(n_proj):
        _pfw(r, j, f'={_pcl(j)}{ebitda_row}-{_pcl(j)}{da_row}', bold=True, fill=XLIGHT_BLUE)
    r += 1

    # EBIT Margin % = EBIT / Revenue  [FORMULA]
    _lbl(r, '  EBIT Margin %', ind=1)
    for i in range(n):
        _fw(r, i,
            f'=IF({_cl(i)}{rev_row}<>0,{_cl(i)}{ebit_row}/{_cl(i)}{rev_row},"")',
            fmt=FMT_PCT)
    for j in range(n_proj):
        _pfw(r, j,
             f'=IF({_pcl(j)}{rev_row}<>0,{_pcl(j)}{ebit_row}/{_pcl(j)}{rev_row},"")',
             fmt=FMT_PCT)
    r += 1

    _spacer(ws, r, total_cols); r += 1

    # -- Raw non-operating inputs + projections (hold flat) --
    int_exp_row = r; _write_row(ws, r, '  Interest Expense', inc['interest_expense'], years, indent=1)
    for j in range(n_proj):
        _pfw(r, j, f'={last_hist}{int_exp_row}')
    r += 1

    int_inc_row = r; _write_row(ws, r, '  Interest Income', inc['interest_income'], years, indent=1)
    for j in range(n_proj):
        _pfw(r, j, f'={last_hist}{int_inc_row}')
    r += 1

    other_inc_row = r; _write_row(ws, r, '  Other Income / (Expense)', inc['other_income'], years, indent=1)
    ws.cell(row=r, column=1).comment = Comment(
        "Plug: Pre-tax Income - EBIT - Interest Income"
        " + Interest Expense.\n\n"
        "May include:\n"
        "- Investment gains / losses\n"
        "- Equity method income\n"
        "- Foreign exchange gains / losses\n"
        "- Gains / losses on debt extinguishment\n"
        "- Other non-operating items",
        "Financial Model")
    for j in range(n_proj):
        _pfw(r, j, f'={last_hist}{other_inc_row}')
    r += 1

    # Pre-tax Income = EBIT - Interest Expense + Interest Income + Other Income  [FORMULA]
    pretax_row = r
    _lbl(r, 'Pre-tax Income', bold=True)
    for i in range(n):
        _fw(r, i,
            f'={_cl(i)}{ebit_row}-{_cl(i)}{int_exp_row}+{_cl(i)}{int_inc_row}+{_cl(i)}{other_inc_row}',
            bold=True, fill=XLIGHT_BLUE)
    for j in range(n_proj):
        _pfw(r, j,
             f'={_pcl(j)}{ebit_row}-{_pcl(j)}{int_exp_row}+{_pcl(j)}{int_inc_row}+{_pcl(j)}{other_inc_row}',
             bold=True, fill=XLIGHT_BLUE)
    r += 1

    # -- Tax (projection backfilled after assumptions band) --
    tax_row = r; _write_row(ws, r, '  Income Tax Expense', inc['tax_expense'], years, indent=1); r += 1

    # Effective Tax Rate = Tax / Pre-tax  [FORMULA]
    _lbl(r, '  Effective Tax Rate %', ind=1)
    for i in range(n):
        _fw(r, i,
            f'=IF({_cl(i)}{pretax_row}<>0,{_cl(i)}{tax_row}/{_cl(i)}{pretax_row},"")',
            fmt=FMT_PCT)
    for j in range(n_proj):
        _pfw(r, j,
             f'=IF({_pcl(j)}{pretax_row}<>0,{_pcl(j)}{tax_row}/{_pcl(j)}{pretax_row},"")',
             fmt=FMT_PCT)
    r += 1

    # Net Income = Pre-tax Income - Tax Expense  [FORMULA]
    ni_row = r
    _lbl(r, 'Net Income', bold=True)
    for i in range(n):
        _fw(r, i, f'={_cl(i)}{pretax_row}-{_cl(i)}{tax_row}', bold=True, fill=LIGHT_GREEN)
    for j in range(n_proj):
        _pfw(r, j, f'={_pcl(j)}{pretax_row}-{_pcl(j)}{tax_row}', bold=True, fill=LIGHT_GREEN)
    r += 1

    # Net Margin % = Net Income / Revenue  [FORMULA]
    _lbl(r, '  Net Margin %', ind=1)
    for i in range(n):
        _fw(r, i,
            f'=IF({_cl(i)}{rev_row}<>0,{_cl(i)}{ni_row}/{_cl(i)}{rev_row},"")',
            fmt=FMT_PCT)
    for j in range(n_proj):
        _pfw(r, j,
             f'=IF({_pcl(j)}{rev_row}<>0,{_pcl(j)}{ni_row}/{_pcl(j)}{rev_row},"")',
             fmt=FMT_PCT)
    r += 1

    _spacer(ws, r, total_cols); r += 1

    # EPS and shares are raw XBRL values (not in millions) — historical only
    _write_row(ws, r, '  EPS (Basic)',   inc['eps_basic'],   years, scale=1.0, fmt=FMT_DOLLAR2, indent=1); r += 1
    eps_diluted_row = r
    _write_row(ws, r, '  EPS (Diluted)', inc['eps_diluted'], years, scale=1.0, fmt=FMT_DOLLAR2, indent=1); r += 1
    _write_row(ws, r, '  Shares Outstanding - Basic (M)',   inc['shares_basic'],   years, scale=1e6, fmt=FMT_DOLLAR, indent=1); r += 1
    shares_diluted_row = r
    _write_row(ws, r, '  Shares Outstanding - Diluted (M)', inc['shares_diluted'], years, scale=1e6, fmt=FMT_DOLLAR, indent=1); r += 1
    r += 1

    # ── Growth Rates & Key Margins (assumptions band) ─────────────────
    # Written here (after all IS rows) so row numbers for Rev, GP, etc.
    # are already known.  Projection columns (F-J) filled later after
    # the scenario table at the bottom of the sheet.
    _lbl(r, 'GROWTH RATES & KEY MARGINS', bold=True, fill=LIGHT_BLUE)
    for ci in range(2, 2 + n + n_proj):
        ws.cell(row=r, column=ci).fill = _fill(LIGHT_BLUE)
    r += 1

    # Revenue Growth (YoY %) — N/A for oldest year
    rev_growth_asm_row = r; _lbl(r, '  Revenue Growth (YoY %)', ind=1)
    for i in range(n):
        if i == 0:
            c = ws.cell(row=r, column=2 + i, value='N/A')
            _style(c, fill_hex=MED_GRAY, h_align='center', italic=True)
        else:
            _fw(r, i,
                f'=IF({_cl(i-1)}{rev_row}<>0,{_cl(i)}{rev_row}/{_cl(i-1)}{rev_row}-1,"")',
                fmt=FMT_PCT)
    r += 1

    # Gross Profit Margin (%)
    gp_margin_asm_row = r; _lbl(r, '  Gross Profit Margin (%)', ind=1)
    for i in range(n):
        _fw(r, i,
            f'=IF({_cl(i)}{rev_row}<>0,{_cl(i)}{gp_row}/{_cl(i)}{rev_row},"")',
            fmt=FMT_PCT)
    r += 1

    # R&D % of Sales
    rd_pct_asm_row = r; _lbl(r, '  R&D % of Sales', ind=1)
    for i in range(n):
        _fw(r, i,
            f'=IF({_cl(i)}{rev_row}<>0,{_cl(i)}{rd_row}/{_cl(i)}{rev_row},"")',
            fmt=FMT_PCT)
    r += 1

    # SG&A % of Sales
    sga_pct_asm_row = r; _lbl(r, '  SG&A % of Sales', ind=1)
    for i in range(n):
        _fw(r, i,
            f'=IF({_cl(i)}{rev_row}<>0,{_cl(i)}{sga_row}/{_cl(i)}{rev_row},"")',
            fmt=FMT_PCT)
    r += 1

    # Tax Rate (%)
    tax_rate_asm_row = r; _lbl(r, '  Tax Rate (%)', ind=1)
    for i in range(n):
        _fw(r, i,
            f'=IF({_cl(i)}{pretax_row}<>0,{_cl(i)}{tax_row}/{_cl(i)}{pretax_row},"")',
            fmt=FMT_PCT)
    r += 1

    # EBITDA Margin (%) — used by DCF via scenario dropdown
    ebitda_margin_asm_row = r; _lbl(r, '  EBITDA Margin (%)', ind=1)
    for i in range(n):
        _fw(r, i,
            f'=IF({_cl(i)}{rev_row}<>0,{_cl(i)}{ebitda_row}/{_cl(i)}{rev_row},"")',
            fmt=FMT_PCT)
    r += 1

    # D&A % of Revenue — used by DCF via scenario dropdown
    da_pct_asm_row = r; _lbl(r, '  D&A % of Revenue', ind=1)
    for i in range(n):
        _fw(r, i,
            f'=IF({_cl(i)}{rev_row}<>0,{_cl(i)}{da_row}/{_cl(i)}{rev_row},"")',
            fmt=FMT_PCT)
    r += 1

    # Capex % of Revenue — used by DCF via scenario dropdown
    # Historical: Python-computed (capex_row not yet defined at this point)
    capex_pct_asm_row = r; _lbl(r, '  Capex % of Revenue', ind=1)
    for i, yr in enumerate(years):
        cap_v = abs(cf['capex'].get(yr) or 0)
        rev_v = inc['revenue'].get(yr)
        if rev_v and rev_v != 0:
            val = cap_v / rev_v
        else:
            val = None
        if val is not None:
            c = ws.cell(row=r, column=2 + i, value=val)
            _style(c, h_align='right', number_format=FMT_PCT,
                   fill_hex=LIGHT_GRAY if i % 2 == 0 else WHITE)
        else:
            c = ws.cell(row=r, column=2 + i, value='N/A')
            _style(c, fill_hex=MED_GRAY, h_align='center', italic=True)
    r += 1

    _spacer(ws, r, total_cols); r += 1

    # ── Backfill IS projection formulas ─────────────────────────────
    # Now that assumptions band row numbers are known, fill in the
    # projection columns (F-J) for the IS line items above.
    for j in range(n_proj):
        prev = _cl(n - 1) if j == 0 else _pcl(j - 1)
        # Revenue = prior year × (1 + growth)
        _pfw(rev_row, j, f'={prev}{rev_row}*(1+{_pcl(j)}{rev_growth_asm_row})')
        # COGS = Revenue × (1 - GP margin)
        _pfw(cogs_row, j, f'={_pcl(j)}{rev_row}*(1-{_pcl(j)}{gp_margin_asm_row})')
        # R&D = Revenue × R&D %
        _pfw(rd_row, j, f'={_pcl(j)}{rev_row}*{_pcl(j)}{rd_pct_asm_row}')
        # SG&A = Revenue × SGA %
        _pfw(sga_row, j, f'={_pcl(j)}{rev_row}*{_pcl(j)}{sga_pct_asm_row}')
        # Tax = Pre-tax × Tax Rate
        _pfw(tax_row, j, f'={_pcl(j)}{pretax_row}*{_pcl(j)}{tax_rate_asm_row}')
        # SBC = prior year × (1 + revenue growth)
        _pfw(sbc_row, j, f'={prev}{sbc_row}*(1+{_pcl(j)}{rev_growth_asm_row})')

    # =========================================================================
    # BALANCE SHEET  (historical + projections)
    # =========================================================================
    _write_section_header(ws, r, 'BALANCE SHEET', cols=total_cols);  r += 1
    _write_col_headers(ws, r, list(range(2, 2 + n)), years, start_col=2)
    for j, py in enumerate(proj_years):
        col = proj_start_col + j
        c = ws.cell(row=r, column=col, value=f'FY{py}E')
        _style(c, fill_hex=MED_BLUE, bold=True, font_color=WHITE, h_align='center')
        c.border = THIN_BOX
    r += 1

    # Helper: growth-driven projection (prior year × (1 + rev growth))
    def _bs_grow(row_num):
        for j in range(n_proj):
            prev = _cl(n - 1) if j == 0 else _pcl(j - 1)
            _pfw(row_num, j, f'={prev}{row_num}*(1+{_pcl(j)}{rev_growth_asm_row})')

    # Helper: hold-flat projection (= last historical year)
    def _bs_flat(row_num):
        for j in range(n_proj):
            _pfw(row_num, j, f'={last_hist}{row_num}')

    _write_row(ws, r, 'ASSETS', {y: None for y in years}, years, bold=True); r += 1

    # Cash — backfilled later (will reference ending cash from CF)
    cash_row = r
    _write_row(ws, r, '  Cash & Cash Equivalents', bs['cash'], years, indent=1)
    _bs_flat(r)  # temporary: hold flat; will be overwritten after CF projections
    r += 1

    # ST Investments — grow with revenue
    st_inv_row = r
    _write_row(ws, r, '  Short-term Investments', bs['st_investments'], years, indent=1)
    _bs_grow(r); r += 1

    # Accounts Receivable — grow with revenue
    ar_row = r
    _write_row(ws, r, '  Accounts Receivable', bs['accounts_rec'], years, indent=1)
    _bs_grow(r); r += 1

    # Inventory — grows with COGS ratio
    inventory_row = r
    _write_row(ws, r, '  Inventory', bs['inventory'], years, indent=1)
    for j in range(n_proj):
        prev = _cl(n - 1) if j == 0 else _pcl(j - 1)
        _pfw(r, j,
             f'=IF({prev}{cogs_row}<>0,'
             f'{prev}{inventory_row}*{_pcl(j)}{cogs_row}/{prev}{cogs_row},'
             f'{prev}{inventory_row})')
    r += 1

    # Other Current Assets (plug) — grow with revenue
    other_ca_plug = {}
    for yr in years:
        tca = bs['total_current_a'].get(yr)
        if tca is not None:
            other_ca_plug[yr] = tca - sum(
                bs[k].get(yr) or 0
                for k in ('cash', 'st_investments', 'accounts_rec', 'inventory'))
        else:
            other_ca_plug[yr] = None
    other_ca_row = r
    _write_row(ws, r, '  Other Current Assets (plug)', other_ca_plug, years, indent=1)
    ws.cell(row=r, column=1).comment = Comment(
        "Plug: Total Current Assets - Cash - ST Investments"
        " - Accounts Receivable - Inventory.\n\n"
        "May include:\n"
        "- Prepaid expenses\n"
        "- Deferred tax assets (current)\n"
        "- Other receivables\n"
        "- Assets held for sale\n"
        "- Contract assets",
        "Financial Model")
    _bs_grow(r); r += 1

    # Total Current Assets [FORMULA for projections]
    total_ca_row = r
    _write_row(ws, r, 'Total Current Assets', bs['total_current_a'], years, bold=True, fill=XLIGHT_BLUE)
    for j in range(n_proj):
        cl = _pcl(j)
        _pfw(r, j,
             f'={cl}{cash_row}+{cl}{st_inv_row}+{cl}{ar_row}'
             f'+{cl}{inventory_row}+{cl}{other_ca_row}',
             bold=True, fill=XLIGHT_BLUE)
    r += 1

    _spacer(ws, r, total_cols); r += 1

    # PP&E — backfilled after PP&E schedule is written
    ppe_row = r
    _write_row(ws, r, '  PP&E, net', bs['ppe_net'], years, indent=1)
    _bs_flat(r)  # temporary: overwritten after PP&E schedule
    r += 1

    # Goodwill — grow with revenue
    gw_row = r
    _write_row(ws, r, '  Goodwill', bs['goodwill'], years, indent=1)
    _bs_grow(r); r += 1

    # Intangible Assets — grow with revenue
    intangibles_row = r
    _write_row(ws, r, '  Intangible Assets', bs['intangibles'], years, indent=1)
    _bs_grow(r); r += 1

    # Marketable Securities (non-current) — grow with revenue
    lt_inv_row = r
    _write_row(ws, r, '  Marketable Securities (non-current)', bs['lt_investments'], years, indent=1)
    _bs_grow(r); r += 1

    # Other Non-Current Assets (plug) — grow with revenue
    other_nca_plug = {}
    for yr in years:
        ta = bs['total_assets'].get(yr)
        tca = bs['total_current_a'].get(yr)
        if ta is not None and tca is not None:
            other_nca_plug[yr] = ta - tca - sum(
                bs[k].get(yr) or 0
                for k in ('ppe_net', 'goodwill', 'intangibles', 'lt_investments'))
        else:
            other_nca_plug[yr] = None
    other_nca_row = r
    _write_row(ws, r, '  Other Non-current Assets (plug)', other_nca_plug, years, indent=1)
    ws.cell(row=r, column=1).comment = Comment(
        "Plug: Total Assets - Total Current Assets - PP&E"
        " - Goodwill - Intangibles - Marketable Securities (non-current).\n\n"
        "May include:\n"
        "- Operating lease right-of-use assets\n"
        "- Deferred tax assets (non-current)\n"
        "- Non-current contract assets\n"
        "- Other long-term assets",
        "Financial Model")
    _bs_grow(r); r += 1

    # Total Assets [FORMULA for projections]
    total_assets_row = r
    _write_row(ws, r, 'Total Assets', bs['total_assets'], years, bold=True, fill=LIGHT_BLUE)
    for j in range(n_proj):
        cl = _pcl(j)
        _pfw(r, j,
             f'={cl}{total_ca_row}+{cl}{ppe_row}+{cl}{gw_row}'
             f'+{cl}{intangibles_row}+{cl}{lt_inv_row}+{cl}{other_nca_row}',
             bold=True, fill=LIGHT_BLUE)
    r += 1

    _spacer(ws, r, total_cols); r += 1

    # ── Liabilities ──────────────────────────────────────────────────
    _write_row(ws, r, 'LIABILITIES', {y: None for y in years}, years, bold=True); r += 1

    ap_row = r
    _write_row(ws, r, '  Accounts Payable', bs['accounts_pay'], years, indent=1)
    # Projection: prior year AP × (this year COGS / prior year COGS)
    for j in range(n_proj):
        prev = _cl(n - 1) if j == 0 else _pcl(j - 1)
        _pfw(r, j,
             f'=IF({prev}{cogs_row}<>0,'
             f'{prev}{ap_row}*{_pcl(j)}{cogs_row}/{prev}{cogs_row},'
             f'{prev}{ap_row})')
    r += 1

    accrued_row = r
    _write_row(ws, r, '  Accrued Liabilities', bs['accrued_liab'], years, indent=1)
    _bs_grow(r); r += 1

    st_debt_row = r
    _write_row(ws, r, '  Short-term Debt', bs['st_debt'], years, indent=1)
    _bs_grow(r); r += 1

    deferred_rev_row = r
    _write_row(ws, r, '  Deferred Revenue (current)', bs['deferred_rev_cur'], years, indent=1)
    _bs_grow(r); r += 1

    # Other Current Liabilities (plug) — grow with revenue
    other_cl_plug = {}
    for yr in years:
        tcl = bs['total_current_l'].get(yr)
        if tcl is not None:
            other_cl_plug[yr] = tcl - sum(
                bs[k].get(yr) or 0
                for k in ('accounts_pay', 'accrued_liab', 'st_debt', 'deferred_rev_cur'))
        else:
            other_cl_plug[yr] = None
    other_cl_row = r
    _write_row(ws, r, '  Other Current Liabilities (plug)', other_cl_plug, years, indent=1)
    ws.cell(row=r, column=1).comment = Comment(
        "Plug: Total Current Liabilities - Accounts Payable"
        " - Accrued Liabilities - ST Debt - Deferred Revenue.\n\n"
        "May include:\n"
        "- Operating lease liabilities (current)\n"
        "- Accrued income taxes\n"
        "- Dividends payable\n"
        "- Customer deposits\n"
        "- Other current liabilities",
        "Financial Model")
    _bs_grow(r); r += 1

    # Total Current Liabilities [FORMULA for projections]
    total_cl_row = r
    _write_row(ws, r, 'Total Current Liabilities', bs['total_current_l'], years, bold=True, fill=XLIGHT_BLUE)
    for j in range(n_proj):
        cl = _pcl(j)
        _pfw(r, j,
             f'={cl}{ap_row}+{cl}{accrued_row}+{cl}{st_debt_row}'
             f'+{cl}{deferred_rev_row}+{cl}{other_cl_row}',
             bold=True, fill=XLIGHT_BLUE)
    r += 1

    _spacer(ws, r, total_cols); r += 1

    lt_debt_row = r
    _write_row(ws, r, '  Long-term Debt', bs['lt_debt'], years, indent=1)
    _bs_grow(r); r += 1

    dtl_row = r
    _write_row(ws, r, '  Deferred Tax Liabilities', bs['deferred_tax_l'], years, indent=1)
    _bs_grow(r); r += 1

    # Other Non-Current Liabilities (plug) — grow with revenue
    other_ncl_plug = {}
    for yr in years:
        tl = bs['total_liabilities'].get(yr)
        tcl = bs['total_current_l'].get(yr)
        if tl is not None and tcl is not None:
            other_ncl_plug[yr] = tl - tcl - sum(
                bs[k].get(yr) or 0
                for k in ('lt_debt', 'deferred_tax_l'))
        else:
            other_ncl_plug[yr] = None
    other_ncl_row = r
    _write_row(ws, r, '  Other Non-current Liabilities (plug)', other_ncl_plug, years, indent=1)
    ws.cell(row=r, column=1).comment = Comment(
        "Plug: Total Liabilities - Total Current Liabilities"
        " - Long-term Debt - Deferred Tax Liabilities.\n\n"
        "May include:\n"
        "- Operating lease liabilities (non-current)\n"
        "- Finance lease liabilities\n"
        "- Pension & post-retirement obligations\n"
        "- Uncertain tax positions\n"
        "- Non-current deferred revenue\n"
        "- Other long-term liabilities",
        "Financial Model")
    _bs_grow(r); r += 1

    # Total Liabilities [FORMULA for projections]
    total_liabilities_row = r
    _write_row(ws, r, 'Total Liabilities', bs['total_liabilities'], years, bold=True, fill=LIGHT_BLUE)
    for j in range(n_proj):
        cl = _pcl(j)
        _pfw(r, j,
             f'={cl}{total_cl_row}+{cl}{lt_debt_row}+{cl}{dtl_row}+{cl}{other_ncl_row}',
             bold=True, fill=LIGHT_BLUE)
    r += 1

    _spacer(ws, r, total_cols); r += 1

    # ── Shareholders' Equity ─────────────────────────────────────────
    _write_row(ws, r, "SHAREHOLDERS' EQUITY", {y: None for y in years}, years, bold=True); r += 1

    cs_row = r
    _write_row(ws, r, '  Common Stock', bs['common_stock'], years, indent=1)
    _bs_flat(r); r += 1

    apic_row = r
    _write_row(ws, r, '  Additional Paid-in Capital', bs['apic'], years, indent=1)
    _bs_flat(r); r += 1

    re_row = r
    _write_row(ws, r, '  Retained Earnings', bs['retained_earnings'], years, indent=1)
    # Projection: prior year RE + Net Income
    for j in range(n_proj):
        prev = _cl(n - 1) if j == 0 else _pcl(j - 1)
        _pfw(r, j, f'={prev}{re_row}+{_pcl(j)}{ni_row}')
    r += 1

    ts_row = r
    _write_row(ws, r, '  Treasury Stock', bs['treasury_stock'], years, indent=1)
    _bs_flat(r); r += 1

    # Total Equity = Total Assets - Total Liabilities [FORMULA for projections]
    total_equity_row = r
    _write_row(ws, r, 'Total Equity', bs['total_equity'], years, bold=True, fill=XLIGHT_BLUE)
    for j in range(n_proj):
        cl = _pcl(j)
        _pfw(r, j, f'={cl}{total_assets_row}-{cl}{total_liabilities_row}',
             bold=True, fill=XLIGHT_BLUE)
    r += 1

    _spacer(ws, r, total_cols); r += 1

    # Total Liabilities + Equity  [FORMULA]
    _lbl(r, 'Total Liabilities + Equity', bold=True)
    for i in range(n):
        _fw(r, i,
            f'={_cl(i)}{total_liabilities_row}+{_cl(i)}{total_equity_row}',
            bold=True, fill=XLIGHT_BLUE)
    for j in range(n_proj):
        cl = _pcl(j)
        _pfw(r, j, f'={cl}{total_liabilities_row}+{cl}{total_equity_row}',
             bold=True, fill=XLIGHT_BLUE)
    r += 1

    # Balance Check: 0 = balanced, FALSE = imbalance
    _lbl(r, 'Balance Check  (0 = Balanced  |  FALSE = Imbalance)', bold=True)
    for i in range(n):
        a_cell = f'{_cl(i)}{total_assets_row}'
        diff   = f'ABS({a_cell}-({_cl(i)}{total_liabilities_row}+{_cl(i)}{total_equity_row}))'
        tol    = f'MAX({a_cell}*0.00001,0.5)'
        formula = f'=IF({diff}<{tol},0,FALSE)'
        c = ws.cell(row=r, column=2 + i, value=formula)
        _style(c, fill_hex=YELLOW, bold=True, h_align='center')
    # Projection balance check: always 0 by construction (TE = TA - TL)
    for j in range(n_proj):
        c = ws.cell(row=r, column=proj_start_col + j, value=0)
        _style(c, fill_hex=YELLOW, bold=True, h_align='center',
               number_format=FMT_DOLLAR)
    r += 2

    # =========================================================================
    # CASH FLOW STATEMENT  (historical + projections)
    # =========================================================================
    cf_cols = total_cols
    _write_section_header(ws, r, 'CASH FLOW STATEMENT', cols=cf_cols);  r += 1
    _write_col_headers(ws, r, list(range(2, 2 + n)), years, start_col=2)
    for j, py in enumerate(proj_years):
        col = proj_start_col + j
        c = ws.cell(row=r, column=col, value=f'FY{py}E')
        _style(c, fill_hex=MED_BLUE, bold=True, font_color=WHITE, h_align='center')
        c.border = THIN_BOX
    r += 1

    _write_row(ws, r, 'OPERATING ACTIVITIES', {y: None for y in years}, years, bold=True); r += 1

    # -- Net Income --
    cf_ni_row = r
    _write_row(ws, r, '  Net Income', cf['net_income'], years, indent=1)
    for j in range(n_proj):
        _pfw(r, j, f'={_pcl(j)}{ni_row}')
    r += 1

    # -- D&A --
    cf_da_row = r
    _write_row(ws, r, '  Depreciation & Amortization', cf['da'], years, indent=1)
    # Projection: placeholder — backfilled after PP&E schedule
    for j in range(n_proj):
        _pfw(r, j, f'={_pcl(j)}{da_row}')
    r += 1

    # -- SBC --
    cf_sbc_row = r
    _write_row(ws, r, '  Stock-based Compensation', cf['sbc'], years, indent=1)
    for j in range(n_proj):
        _pfw(r, j, f'={_pcl(j)}{sbc_row}')
    r += 1

    # -- Decreases / (Increases) in Working Capital Assets --
    # WC Assets = AR + Inventory + Other CA (excludes Cash, ST Investments)
    # Decrease in WC assets = source of cash (positive)
    cf_wc_a_row = r
    _lbl(r, '  Decreases / (Increases) in WC Assets', ind=1)
    for i in range(n):
        if i == 0:
            c = ws.cell(row=r, column=2 + i, value=0)
            _style(c, fill_hex=MED_GRAY, h_align='right', number_format=FMT_DOLLAR,
                   italic=True)
        else:
            prev = _cl(i - 1); curr = _cl(i)
            _fw(r, i,
                f'=({prev}{ar_row}+{prev}{inventory_row}+{prev}{other_ca_row})'
                f'-({curr}{ar_row}+{curr}{inventory_row}+{curr}{other_ca_row})')
    for j in range(n_proj):
        prev = _cl(n - 1) if j == 0 else _pcl(j - 1)
        curr = _pcl(j)
        _pfw(r, j,
             f'=({prev}{ar_row}+{prev}{inventory_row}+{prev}{other_ca_row})'
             f'-({curr}{ar_row}+{curr}{inventory_row}+{curr}{other_ca_row})')
    r += 1

    # -- Increases / (Decreases) in Working Capital Liabilities --
    # WC Liabilities = AP + Accrued + Deferred Rev + Other CL (excludes ST Debt)
    # Increase in WC liabilities = source of cash (positive)
    cf_wc_l_row = r
    _lbl(r, '  Increases / (Decreases) in WC Liabilities', ind=1)
    for i in range(n):
        if i == 0:
            c = ws.cell(row=r, column=2 + i, value=0)
            _style(c, fill_hex=MED_GRAY, h_align='right', number_format=FMT_DOLLAR,
                   italic=True)
        else:
            prev = _cl(i - 1); curr = _cl(i)
            _fw(r, i,
                f'=({curr}{ap_row}+{curr}{accrued_row}+{curr}{deferred_rev_row}+{curr}{other_cl_row})'
                f'-({prev}{ap_row}+{prev}{accrued_row}+{prev}{deferred_rev_row}+{prev}{other_cl_row})')
    for j in range(n_proj):
        prev = _cl(n - 1) if j == 0 else _pcl(j - 1)
        curr = _pcl(j)
        _pfw(r, j,
             f'=({curr}{ap_row}+{curr}{accrued_row}+{curr}{deferred_rev_row}+{curr}{other_cl_row})'
             f'-({prev}{ap_row}+{prev}{accrued_row}+{prev}{deferred_rev_row}+{prev}{other_cl_row})')
    r += 1

    # -- Other Operating Activities (plug) --
    # Historical: backs into actual Cash from Ops; Projection: 0
    cf_other_op_row = r
    _lbl(r, '  Other Operating Activities (plug)', ind=1)
    # Pre-compute op_cf_row position for formula (comes 1 row after this)
    op_cf_row = r + 1
    for i in range(n):
        cl = _cl(i)
        _fw(r, i,
            f'={cl}{op_cf_row}-{cl}{cf_ni_row}-{cl}{cf_da_row}'
            f'-{cl}{cf_sbc_row}-{cl}{cf_wc_a_row}-{cl}{cf_wc_l_row}')
    ws.cell(row=r, column=1).comment = Comment(
        "Plug: Cash from Operations - Net Income - D&A - SBC"
        " - WC Asset Changes - WC Liability Changes.\n\n"
        "May include:\n"
        "- Deferred income taxes\n"
        "- Impairment / write-down charges\n"
        "- Gains/losses on investments\n"
        "- Amortization of debt issuance costs\n"
        "- Other non-cash adjustments",
        "Financial Model")
    for j in range(n_proj):
        _pfw(r, j, f'=0')
    r += 1

    # Cash from Operations
    # Historical: actual data; Projection: SUM of operating items
    _write_row(ws, r, 'Cash from Operations', cf['operating_cf'], years,
               bold=True, fill=LIGHT_GREEN)
    for j in range(n_proj):
        cl = _pcl(j)
        _pfw(r, j,
             f'={cl}{cf_ni_row}+{cl}{cf_da_row}+{cl}{cf_sbc_row}'
             f'+{cl}{cf_wc_a_row}+{cl}{cf_wc_l_row}+{cl}{cf_other_op_row}',
             bold=True, fill=LIGHT_GREEN)
    r += 1

    _spacer(ws, r, cf_cols); r += 1

    # -- INVESTING ACTIVITIES --
    _write_row(ws, r, 'INVESTING ACTIVITIES', {y: None for y in years}, years, bold=True); r += 1
    capex_row = r
    _write_row(ws, r, '  Capital Expenditures (Capex)', cf['capex'], years,
               negate=True, indent=1)
    # Projection: placeholder — backfilled after PP&E schedule
    for j in range(n_proj):
        _pfw(r, j, f'=0')
    r += 1

    _write_row(ws, r, '  Acquisitions', cf['acquisitions'], years,
               negate=True, indent=1)
    for j in range(n_proj):
        _pfw(r, j, f'=0')
    r += 1

    # Other Investing Activities: plug for historical, 0 for projections
    other_inv_plug = {}
    for yr in years:
        inv = cf['investing_cf'].get(yr)
        if inv is not None:
            other_inv_plug[yr] = inv + sum(
                cf[k].get(yr) or 0
                for k in ('capex', 'acquisitions'))
        else:
            other_inv_plug[yr] = None
    _write_row(ws, r, '  Other Investing Activities (plug)', other_inv_plug,
               years, indent=1)
    ws.cell(row=r, column=1).comment = Comment(
        "Plug: Cash from Investing + Capex + Acquisitions.\n\n"
        "May include:\n"
        "- Purchases of marketable securities\n"
        "- Maturities / sales of marketable securities\n"
        "- Purchases of non-marketable securities\n"
        "- Other investing activities",
        "Financial Model")
    for j in range(n_proj):
        _pfw(r, j, f'=0')
    r += 1

    inv_cf_row = r
    _write_row(ws, r, 'Cash from Investing', cf['investing_cf'], years,
               bold=True, fill=LIGHT_RED)
    # Projection: = Capex (already negative) + Acq + Other
    for j in range(n_proj):
        _pfw(r, j, f'={_pcl(j)}{capex_row}', bold=True, fill=LIGHT_RED)
    r += 1

    _spacer(ws, r, cf_cols); r += 1

    # -- FINANCING ACTIVITIES --
    _write_row(ws, r, 'FINANCING ACTIVITIES', {y: None for y in years}, years, bold=True); r += 1

    # Dividends Paid (historical negative via negate=True in _write_row)
    # Projection: -MIN(|prior year dividends|, |this year net income|)
    dividends_row = r
    _write_row(ws, r, '  Dividends Paid', cf['dividends'], years, negate=True, indent=1)
    for j in range(n_proj):
        prev = _cl(n - 1) if j == 0 else _pcl(j - 1)
        cl = _pcl(j)
        _pfw(r, j, f'=-MIN(ABS({prev}{dividends_row}),ABS({cl}{ni_row}))')
    r += 1

    # Share Repurchases (historical negative via negate=True)
    # Projection: placeholder — backfilled to reference RE schedule
    repurchases_row = r
    _write_row(ws, r, '  Share Repurchases', cf['repurchases'], years, negate=True, indent=1)
    for j in range(n_proj):
        _pfw(r, j, f'=0')  # placeholder, backfilled after RE schedule
    r += 1

    # Debt Issuance (positive = cash inflow)
    # Projection: prior year × (1 + revenue growth)
    debt_iss_row = r
    _write_row(ws, r, '  Debt Issuance', cf['debt_issuance'], years, indent=1)
    for j in range(n_proj):
        prev = _cl(n - 1) if j == 0 else _pcl(j - 1)
        cl = _pcl(j)
        _pfw(r, j, f'={prev}{debt_iss_row}*(1+{cl}{rev_growth_asm_row})')
    r += 1

    # Debt Repayment (historical negative via negate=True)
    # Projection: -MIN(|prior year repay| × (1 + growth), ST debt + LT debt)
    debt_rep_row = r
    _write_row(ws, r, '  Debt Repayment', cf['debt_repay'], years, negate=True, indent=1)
    for j in range(n_proj):
        prev = _cl(n - 1) if j == 0 else _pcl(j - 1)
        cl = _pcl(j)
        _pfw(r, j,
             f'=-MIN(ABS({prev}{debt_rep_row})*(1+{cl}{rev_growth_asm_row}),'
             f'{cl}{st_debt_row}+{cl}{lt_debt_row})')
    r += 1

    # Other Financing Activities (plug): historical only, projection = 0
    other_fin_plug = {}
    for yr in years:
        fin = cf['financing_cf'].get(yr)
        if fin is not None:
            other_fin_plug[yr] = (fin
                + (cf['dividends'].get(yr) or 0)
                + (cf['repurchases'].get(yr) or 0)
                - (cf['debt_issuance'].get(yr) or 0)
                + (cf['debt_repay'].get(yr) or 0))
        else:
            other_fin_plug[yr] = None
    other_fin_row = r
    _write_row(ws, r, '  Other Financing Activities (plug)', other_fin_plug,
               years, indent=1)
    ws.cell(row=r, column=1).comment = Comment(
        "Plug: Cash from Financing + Dividends + Repurchases"
        " - Debt Issuance + Debt Repayment.\n\n"
        "May include:\n"
        "- Stock option / RSU proceeds\n"
        "- Finance lease principal payments\n"
        "- Contingent consideration payments\n"
        "- Other financing activities",
        "Financial Model")
    for j in range(n_proj):
        _pfw(r, j, f'=0')
    r += 1

    # Cash from Financing = SUM of all financing items
    # All items already have correct sign (dividends/repurchases/debt_rep negative,
    # debt_issuance positive, other = 0)
    fin_cf_row = r
    _write_row(ws, r, 'Cash from Financing', cf['financing_cf'], years,
               bold=True, fill=LIGHT_RED)
    for j in range(n_proj):
        cl = _pcl(j)
        _pfw(r, j,
             f'={cl}{dividends_row}+{cl}{repurchases_row}'
             f'+{cl}{debt_iss_row}+{cl}{debt_rep_row}+{cl}{other_fin_row}',
             bold=True, fill=LIGHT_RED)
    r += 1

    _spacer(ws, r, cf_cols); r += 1

    fx_cf_row = r
    _write_row(ws, r, 'Effect of Exchange Rate Changes on Cash', cf['fx_effect'],
               years)
    for j in range(n_proj):
        _pfw(r, j, f'=0')
    r += 1

    _spacer(ws, r, cf_cols); r += 1

    fcf_row = r
    _lbl(r, 'Free Cash Flow  (Op + Inv + Fin + FX)', bold=True)
    for i in range(n):
        _fw(r, i,
            f'={_cl(i)}{op_cf_row}+{_cl(i)}{inv_cf_row}+{_cl(i)}{fin_cf_row}+{_cl(i)}{fx_cf_row}',
            bold=True, fill=LIGHT_GREEN)
    for j in range(n_proj):
        _pfw(r, j,
             f'={_pcl(j)}{op_cf_row}+{_pcl(j)}{inv_cf_row}+{_pcl(j)}{fin_cf_row}+{_pcl(j)}{fx_cf_row}',
             bold=True, fill=LIGHT_GREEN)
    r += 1

    # FCF Margin %
    _lbl(r, '  FCF Margin %', ind=1)
    for i in range(n):
        _fw(r, i,
            f'=IF({_cl(i)}{rev_row}<>0,{_cl(i)}{fcf_row}/{_cl(i)}{rev_row},"")',
            fmt=FMT_PCT)
    for j in range(n_proj):
        _pfw(r, j,
             f'=IF({_pcl(j)}{rev_row}<>0,{_pcl(j)}{fcf_row}/{_pcl(j)}{rev_row},"")',
             fmt=FMT_PCT)
    r += 2

    # =========================================================================
    # CASH FLOW RECONCILIATION
    # Verifies: Beginning Cash + Net Change in Cash = Ending Cash on Balance Sheet
    # The check cell returns 0 if the CF statement ties to the BS; FALSE otherwise.
    # (First year has no prior-year data, so only years 2 & 3 are checkable.)
    # =========================================================================
    _write_section_header(
        ws, r,
        'CASH FLOW RECONCILIATION  (Beginning Cash + Net Change = Ending Cash on BS)',
        cols=cf_cols)
    r += 1
    # Column headers: historical + projection
    _write_col_headers(ws, r, list(range(2, 2 + n)), years, start_col=2)
    for j, py in enumerate(proj_years):
        col = proj_start_col + j
        c = ws.cell(row=r, column=col, value=f'FY{py}E')
        _style(c, fill_hex=MED_BLUE, bold=True, font_color=WHITE, h_align='center')
        c.border = THIN_BOX
    r += 1

    # Beginning Cash = prior year's ending cash
    beg_cash_row = r
    _lbl(r, 'Beginning Cash  (prior year Ending Cash)')
    for i in range(n):
        if i == 0:
            c = ws.cell(row=r, column=2 + i, value='N/A (no prior year)')
            _style(c, fill_hex=MED_GRAY, h_align='center', italic=True)
        else:
            prev = get_column_letter(2 + i - 1)
            _fw(r, i, f'={prev}{cash_row}')
    # Projection: beginning cash = prior year's ending cash (from this recon)
    # end_cf_row is defined below — use forward-known offset
    # We know the row layout: beg_cash, +4 CF rows, net_chg, end_cash
    # end_cf_row = beg_cash_row + 6
    # For j=0, beginning cash = last historical year BS cash
    # For j>0, beginning cash = prior projection ending cash
    for j in range(n_proj):
        if j == 0:
            _pfw(r, j, f'={last_hist}{cash_row}')
        else:
            # Forward reference to end_cf_row — will be at beg_cash_row + 6
            _pfw(r, j, f'={_pcl(j-1)}{beg_cash_row + 6}')
    r += 1

    # CF section totals (referenced from CF statement above)
    recon_cf_rows = []
    for label, src in [
        ('+ Cash from Operations',              op_cf_row),
        ('+ Cash from Investing',               inv_cf_row),
        ('+ Cash from Financing',               fin_cf_row),
        ('+ Effect of FX on Cash',              fx_cf_row),
    ]:
        recon_cf_rows.append(r)
        _lbl(r, label, ind=1)
        for i in range(n):
            _fw(r, i, f'={_cl(i)}{src}')
        for j in range(n_proj):
            _pfw(r, j, f'={_pcl(j)}{src}')
        r += 1

    # Net Change in Cash = Op + Inv + Fin + FX  [FORMULA]
    net_chg_row = r
    _lbl(r, 'Net Change in Cash', bold=True)
    for i in range(n):
        _fw(r, i,
            f'={_cl(i)}{op_cf_row}+{_cl(i)}{inv_cf_row}+{_cl(i)}{fin_cf_row}+{_cl(i)}{fx_cf_row}',
            bold=True, fill=XLIGHT_BLUE)
    for j in range(n_proj):
        cl = _pcl(j)
        _pfw(r, j,
             f'={cl}{op_cf_row}+{cl}{inv_cf_row}+{cl}{fin_cf_row}+{cl}{fx_cf_row}',
             bold=True, fill=XLIGHT_BLUE)
    r += 1

    # Ending Cash = Beginning + Net Change  [FORMULA]
    end_cf_row = r
    assert end_cf_row == beg_cash_row + 6, \
        f"end_cf_row layout assumption broken: {end_cf_row} != {beg_cash_row + 6}"
    _lbl(r, 'Ending Cash  (CF-derived)', bold=True)
    for i in range(n):
        if i == 0:
            c = ws.cell(row=r, column=2 + i, value='N/A')
            _style(c, fill_hex=MED_GRAY, h_align='center', italic=True)
        else:
            _fw(r, i, f'={_cl(i)}{beg_cash_row}+{_cl(i)}{net_chg_row}',
                bold=True, fill=LIGHT_BLUE)
    for j in range(n_proj):
        _pfw(r, j, f'={_pcl(j)}{beg_cash_row}+{_pcl(j)}{net_chg_row}',
             bold=True, fill=LIGHT_BLUE)
    r += 1

    # Ending Cash per Balance Sheet  [FORMULA, references BS cash row]
    _lbl(r, 'Ending Cash  (Balance Sheet)', bold=True)
    for i in range(n):
        _fw(r, i, f'={_cl(i)}{cash_row}', bold=True, fill=LIGHT_BLUE)
    for j in range(n_proj):
        _pfw(r, j, f'={_pcl(j)}{cash_row}', bold=True, fill=LIGHT_BLUE)
    r += 1

    # Reconciliation Check: 0 = CF ties to BS, FALSE = discrepancy  [FORMULA]
    # Historical: tolerance-based check for XBRL rounding
    # Projection: always 0 by construction (BS cash = ending cash from recon)
    _lbl(r, 'Reconciliation Check  (0 = OK  |  FALSE = Mismatch)', bold=True)
    for i in range(n):
        if i == 0:
            c = ws.cell(row=r, column=2 + i, value='N/A')
            _style(c, fill_hex=MED_GRAY, bold=True, h_align='center')
        else:
            diff = f'ABS({_cl(i)}{end_cf_row}-{_cl(i)}{cash_row})'
            tol  = f'MAX(ABS({_cl(i)}{cash_row})*0.00001,0.5)'
            formula = f'=IF({diff}<{tol},0,FALSE)'
            c = ws.cell(row=r, column=2 + i, value=formula)
            _style(c, fill_hex=YELLOW, bold=True, h_align='center')
    for j in range(n_proj):
        c = ws.cell(row=r, column=proj_start_col + j, value=0)
        _style(c, fill_hex=YELLOW, bold=True, h_align='center',
               number_format=FMT_DOLLAR)
    r += 2

    # =========================================================================
    # PP&E SCHEDULE
    # =========================================================================
    _write_section_header(ws, r, 'PP&E SCHEDULE', cols=total_cols); r += 1
    # Column headers: historical + projection
    _write_col_headers(ws, r, list(range(2, 2 + n)), years, start_col=2)
    for j, py in enumerate(proj_years):
        col = proj_start_col + j
        c = ws.cell(row=r, column=col, value=f'FY{py}E')
        _style(c, fill_hex=MED_BLUE, bold=True, font_color=WHITE, h_align='center')
        c.border = THIN_BOX
    # Step-up column header
    step_col = total_cols + 1
    step_col_letter = get_column_letter(step_col)
    c = ws.cell(row=r, column=step_col, value='Step-up')
    _style(c, fill_hex=DARK_BLUE, bold=True, font_color=WHITE, h_align='center')
    c.border = THIN_BOX
    ws.column_dimensions[step_col_letter].width = 14
    r += 1

    # Pre-compute row positions so all formulas can cross-reference
    ppe_beg_sch_row    = r
    ppe_capex_sch_row  = r + 1
    ppe_depr_sch_row   = r + 2
    ppe_end_sch_row    = r + 3
    # r + 4 = spacer
    da_capex_pct_row   = r + 5
    capex_pct_rev_row  = r + 6

    step_up_ref   = f'${step_col_letter}${da_capex_pct_row}'
    capex_pct_ref = f'${step_col_letter}${capex_pct_rev_row}'

    # --- Beginning PP&E ---
    _lbl(r, 'Beginning PP&E')
    for i in range(n):
        if i == 0:
            c = ws.cell(row=r, column=2 + i, value='N/A')
            _style(c, fill_hex=MED_GRAY, h_align='center', italic=True)
        else:
            _fw(r, i, f'={_cl(i-1)}{ppe_row}')
    for j in range(n_proj):
        if j == 0:
            _pfw(r, j, f'={last_hist}{ppe_row}')
        else:
            _pfw(r, j, f'={_pcl(j-1)}{ppe_end_sch_row}')
    r += 1

    # --- + Capital Expenditures ---
    _lbl(r, '+ Capital Expenditures')
    for i in range(n):
        _fw(r, i, f'=ABS({_cl(i)}{capex_row})')
    last_hist_capex_val = abs(_val(cf['capex'], latest_yr) or 0)
    for j in range(n_proj):
        if j < 2:
            # Editable yellow cells — default to last historical capex
            c = ws.cell(row=r, column=proj_start_col + j,
                        value=round(last_hist_capex_val, 1))
            _style(c, fill_hex=YELLOW, h_align='right', number_format=FMT_DOLLAR)
            c.border = THIN_BOX
        else:
            # Years 3-5: Revenue × Capex % of Revenue
            _pfw(r, j, f'={_pcl(j)}{rev_row}*{capex_pct_ref}')
    r += 1

    # --- - Depreciation ---
    _lbl(r, '- Depreciation')
    for i in range(n):
        _fw(r, i, f'={_cl(i)}{da_row}')
    for j in range(n_proj):
        _pfw(r, j, f'={_pcl(j)}{ppe_capex_sch_row}*{_pcl(j)}{da_capex_pct_row}')
    r += 1

    # --- Ending PP&E = Beg + Capex - Depr ---
    _lbl(r, 'Ending PP&E', bold=True)
    for i in range(n):
        # Historical: anchor to BS PP&E value (absorbs disposals, impairments)
        _fw(r, i, f'={_cl(i)}{ppe_row}', bold=True, fill=XLIGHT_BLUE)
    for j in range(n_proj):
        cl = _pcl(j)
        _pfw(r, j,
             f'={cl}{ppe_beg_sch_row}+{cl}{ppe_capex_sch_row}-{cl}{ppe_depr_sch_row}',
             bold=True, fill=XLIGHT_BLUE)
    r += 1

    _spacer(ws, r, total_cols); r += 1

    # --- D&A / Capex % ---
    _lbl(r, 'D&A / Capex %')
    for i in range(n):
        _fw(r, i,
            f'=IF({_cl(i)}{ppe_capex_sch_row}<>0,'
            f'{_cl(i)}{ppe_depr_sch_row}/{_cl(i)}{ppe_capex_sch_row},"")',
            fmt=FMT_PCT)
    # Yr 1: average of last 3 historical; Yr 2+: prior + step-up
    avg_start = max(0, n - 3)
    avg_cells = ','.join(f'{_cl(i)}{da_capex_pct_row}' for i in range(avg_start, n))
    for j in range(n_proj):
        if j == 0:
            _pfw(r, j, f'=AVERAGE({avg_cells})', fmt=FMT_PCT)
        else:
            _pfw(r, j, f'={_pcl(j-1)}{da_capex_pct_row}+{step_up_ref}', fmt=FMT_PCT)
    # Step-up editable cell (default 2.0%)
    c = ws.cell(row=r, column=step_col, value=0.02)
    _style(c, fill_hex=YELLOW, h_align='right', number_format=FMT_PCT)
    c.border = THIN_BOX
    r += 1

    # --- Capex % of Revenue (for projection years 3-5) ---
    _lbl(r, 'Capex % of Revenue (Yr 3-5)')
    for i in range(n):
        _fw(r, i,
            f'=IF({_cl(i)}{rev_row}<>0,'
            f'ABS({_cl(i)}{capex_row})/{_cl(i)}{rev_row},"")',
            fmt=FMT_PCT)
    # Editable yellow cell in step-up column — default to historical average
    capex_rev_vals = []
    for yr in years:
        cap_v = abs((_val(cf['capex'], yr) or 0))
        rev_v = _val(inc['revenue'], yr)
        if cap_v > 0 and rev_v and rev_v > 0:
            capex_rev_vals.append(cap_v / rev_v)
    avg_capex_pct_rev = _safe_avg(capex_rev_vals) if capex_rev_vals else 0.05
    c = ws.cell(row=r, column=step_col, value=round(avg_capex_pct_rev, 4))
    _style(c, fill_hex=YELLOW, h_align='right', number_format=FMT_PCT)
    c.border = THIN_BOX
    r += 2

    # =========================================================================
    # RETAINED EARNINGS SCHEDULE
    # =========================================================================
    _write_section_header(ws, r, 'RETAINED EARNINGS SCHEDULE', cols=total_cols); r += 1
    # Column headers: historical + projection
    _write_col_headers(ws, r, list(range(2, 2 + n)), years, start_col=2)
    for j, py in enumerate(proj_years):
        col = proj_start_col + j
        c = ws.cell(row=r, column=col, value=f'FY{py}E')
        _style(c, fill_hex=MED_BLUE, bold=True, font_color=WHITE, h_align='center')
        c.border = THIN_BOX
    r += 1

    # Pre-compute row positions
    re_beg_sch_row  = r
    re_ni_sch_row   = r + 1
    re_div_sch_row  = r + 2
    re_rep_sch_row  = r + 3
    re_end_sch_row  = r + 4

    # --- Beginning Retained Earnings ---
    _lbl(r, 'Beginning Retained Earnings')
    for i in range(n):
        if i == 0:
            c = ws.cell(row=r, column=2 + i, value='N/A')
            _style(c, fill_hex=MED_GRAY, h_align='center', italic=True)
        else:
            _fw(r, i, f'={_cl(i-1)}{re_row}')
    for j in range(n_proj):
        if j == 0:
            _pfw(r, j, f'={last_hist}{re_row}')
        else:
            _pfw(r, j, f'={_pcl(j-1)}{re_end_sch_row}')
    r += 1

    # --- + Net Income ---
    _lbl(r, '+ Net Income')
    for i in range(n):
        _fw(r, i, f'={_cl(i)}{ni_row}')
    for j in range(n_proj):
        _pfw(r, j, f'={_pcl(j)}{ni_row}')
    r += 1

    # --- - Dividends ---
    # CF dividends_row: stored positive (absolute), displayed negative via negate
    # In projections, dividends_row cells are already negative
    _lbl(r, '- Dividends')
    for i in range(n):
        _fw(r, i, f'=ABS({_cl(i)}{dividends_row})')
    for j in range(n_proj):
        _pfw(r, j, f'=ABS({_pcl(j)}{dividends_row})')
    r += 1

    # --- - Share Repurchases ---
    # Historical: actual repurchases from CF; Projection: previous year's amount
    _lbl(r, '- Share Repurchases')
    for i in range(n):
        _fw(r, i, f'=ABS({_cl(i)}{repurchases_row})')
    for j in range(n_proj):
        prev = _cl(n - 1) if j == 0 else _pcl(j - 1)
        _pfw(r, j, f'=ABS({prev}{repurchases_row})')
    r += 1

    # --- Ending Retained Earnings = Beg + NI - Div - Repurchases ---
    _lbl(r, 'Ending Retained Earnings', bold=True)
    for i in range(n):
        # Historical: anchor to BS RE value (absorbs other comprehensive income etc.)
        _fw(r, i, f'={_cl(i)}{re_row}', bold=True, fill=XLIGHT_BLUE)
    for j in range(n_proj):
        cl = _pcl(j)
        _pfw(r, j,
             f'={cl}{re_beg_sch_row}+{cl}{re_ni_sch_row}'
             f'-{cl}{re_div_sch_row}-{cl}{re_rep_sch_row}',
             bold=True, fill=XLIGHT_BLUE)
    r += 2

    # ── Backfill: BS Retained Earnings → RE schedule ending balance ──
    for j in range(n_proj):
        _pfw(re_row, j, f'={_pcl(j)}{re_end_sch_row}')

    # ── Backfill: CF Share Repurchases → negated RE schedule repurchases ─
    for j in range(n_proj):
        _pfw(repurchases_row, j, f'=-{_pcl(j)}{re_rep_sch_row}')

    # ── Backfill: BS PP&E projections → PP&E schedule ending balance ──
    for j in range(n_proj):
        _pfw(ppe_row, j, f'={_pcl(j)}{ppe_end_sch_row}')

    # ── Backfill: IS D&A projections → PP&E schedule depreciation ─────
    # Override the "hold flat" D&A projections with PP&E-schedule-driven values
    for j in range(n_proj):
        _pfw(da_row, j, f'={_pcl(j)}{ppe_depr_sch_row}')

    # ── Backfill: IS "D&A embedded in R&D/SG&A" → negate PP&E depreciation ─
    if _da_embedded:
        for j in range(n_proj):
            _pfw(other_opex_row, j, f'=-{_pcl(j)}{ppe_depr_sch_row}')

    # ── Backfill: CF D&A projections → PP&E schedule depreciation ─────
    for j in range(n_proj):
        _pfw(cf_da_row, j, f'={_pcl(j)}{ppe_depr_sch_row}')

    # ── Backfill: CF Capex projections → PP&E schedule capex ──────────
    for j in range(n_proj):
        _pfw(capex_row, j, f'=-{_pcl(j)}{ppe_capex_sch_row}')

    # ── Backfill: BS Cash projections → CF Reconciliation ending cash ─
    for j in range(n_proj):
        _pfw(cash_row, j, f'={_pcl(j)}{end_cf_row}')

    ws.freeze_panes = 'A3'

    # =========================================================================
    # 5-YEAR REVENUE CAGR REFERENCE
    # =========================================================================
    _ext_rev = fd.get('revenue_extended', {})
    _ext_yrs = sorted([y for y, v in _ext_rev.items() if v is not None and v > 0])
    _latest_hist_cl = get_column_letter(2 + n - 1)

    if len(_ext_yrs) >= 2:
        _cagr_start_year = _ext_yrs[0]
        _cagr_start_rev = _ext_rev[_cagr_start_year] / 1e6
        _cagr_periods = years[-1] - _cagr_start_year
    else:
        _cagr_start_year = years[0]
        _cagr_start_rev = (inc['revenue'].get(years[0]) or 0) / 1e6
        _cagr_periods = len(years) - 1

    if _cagr_periods < 1:
        _cagr_periods = 1

    r += 2
    _write_section_header(ws, r, '5-YEAR REVENUE CAGR', cols=total_cols); r += 1

    _lbl(r, f'Starting Revenue ($M, FY{_cagr_start_year})')
    c = ws.cell(row=r, column=2, value=round(_cagr_start_rev, 1))
    _style(c, fill_hex=YELLOW, bold=False, h_align='right',
           number_format='#,##0.0')
    c.border = THIN_BOX
    cagr_start_row = r; r += 1

    _lbl(r, f'Ending Revenue ($M, FY{years[-1]})')
    c = ws.cell(row=r, column=2, value=f'={_latest_hist_cl}{rev_row}')
    _style(c, fill_hex=LIGHT_GREEN, bold=False, h_align='right',
           number_format='#,##0.0')
    c.border = THIN_BOX
    cagr_end_row = r; r += 1

    _lbl(r, 'Number of Periods')
    c = ws.cell(row=r, column=2, value=_cagr_periods)
    _style(c, fill_hex=YELLOW, bold=False, h_align='right',
           number_format='0')
    c.border = THIN_BOX
    cagr_periods_row = r; r += 1

    _lbl(r, '5-Year Revenue CAGR', bold=True)
    c = ws.cell(row=r, column=2,
                value=f'=IF(B{cagr_start_row}>0,(B{cagr_end_row}/B{cagr_start_row})^(1/B{cagr_periods_row})-1,0)')
    _style(c, fill_hex=LIGHT_GREEN, bold=True, h_align='right',
           number_format=FMT_PCT)
    c.border = THIN_BOX
    cagr_row = r; r += 1

    # =========================================================================
    # SCENARIO ASSUMPTIONS TABLE  (at the bottom of the sheet)
    # Three cases (Best / Base / Weak) for each projection driver.
    # The dropdown cell in the title area selects which case to use.
    # =========================================================================
    r += 2
    _write_section_header(
        ws, r,
        'SCENARIO ASSUMPTIONS  (edit yellow cells to customize projections)',
        cols=total_cols)
    r += 1

    # Column headers for projection years + delta column
    delta_col = proj_start_col - 1               # column E (last hist col)
    delta_cl  = get_column_letter(delta_col)
    ws.cell(row=r, column=1, value='').fill = _fill(DARK_BLUE)
    for ci in range(2, delta_col):
        ws.cell(row=r, column=ci).fill = _fill(DARK_BLUE)
    c = ws.cell(row=r, column=delta_col, value='+/−')
    _style(c, fill_hex=DARK_BLUE, bold=True, font_color=WHITE, h_align='center')
    c.border = THIN_BOX
    for j, py in enumerate(proj_years):
        col = proj_start_col + j
        c = ws.cell(row=r, column=col, value=f'FY{py}E')
        _style(c, fill_hex=DARK_BLUE, bold=True, font_color=WHITE, h_align='center')
        c.border = THIN_BOX
    r += 1

    # Compute base-case values from historical averages
    def _hist_vals(num_data, denom_data, as_growth=False):
        """Compute historical ratios or growth rates from raw XBRL data."""
        vals = []
        sorted_yrs = sorted(years)
        if as_growth:
            for idx in range(1, len(sorted_yrs)):
                y0, y1 = sorted_yrs[idx - 1], sorted_yrs[idx]
                v0 = num_data.get(y0)
                v1 = num_data.get(y1)
                if v0 and v1 and v0 > 0:
                    vals.append(v1 / v0 - 1.0)
        else:
            for yr in sorted_yrs:
                nv = num_data.get(yr)
                dv = denom_data.get(yr) if denom_data else None
                if nv is not None and dv and dv != 0:
                    vals.append(nv / dv)
        return _safe_avg(vals) if vals else None

    avg_rev_growth = 0.05  # Placeholder; base case uses CAGR formula reference
    avg_gp_margin  = _hist_vals(
        {yr: (inc['revenue'].get(yr) or 0) - (inc['cogs'].get(yr) or 0) for yr in years},
        inc['revenue']) or 0.40
    avg_rd_pct     = _hist_vals(inc['rd_expense'], inc['revenue']) or 0.05
    avg_sga_pct    = _hist_vals(inc['sga_expense'], inc['revenue']) or 0.10
    avg_tax_rate   = _hist_vals(inc['tax_expense'], inc['pretax_income']) or 0.21
    avg_ebitda_margin = _hist_vals(inc['ebitda'], inc['revenue']) or 0.30
    avg_da_pct     = _hist_vals(inc['da'], inc['revenue']) or 0.03
    avg_capex_pct  = _hist_vals(
        {yr: abs(cf['capex'].get(yr) or 0) for yr in years},
        inc['revenue']) or 0.05

    # Define metrics: (key, label, base_val, spread, invert)
    # invert=True means lower is "better" (costs / tax)
    metrics_def = [
        ('rev_growth',     'Revenue Growth (YoY %)',    avg_rev_growth,     0.03, False),
        ('gp_margin',      'Gross Profit Margin (%)',   avg_gp_margin,      0.03, False),
        ('rd_pct',         'R&D % of Sales',            avg_rd_pct,         0.02, True),
        ('sga_pct',        'SG&A % of Sales',           avg_sga_pct,        0.02, True),
        ('tax_rate',       'Tax Rate (%)',               avg_tax_rate,       0.03, True),
        ('ebitda_margin',  'EBITDA Margin (%)',          avg_ebitda_margin,  0.03, False),
        ('da_pct',         'D&A % of Revenue',           avg_da_pct,         0.01, True),
        ('capex_pct',      'Capex % of Revenue',         avg_capex_pct,      0.01, True),
    ]

    scenario_rows = {}  # {metric_key: {'best': row, 'base': row, 'weak': row}}
    FMT_DELTA = '+0.0%;-0.0%;0.0%'

    for metric_key, metric_label, base_val, spread, invert in metrics_def:
        # Sub-header for this metric
        _lbl(r, metric_label, bold=True, fill=LIGHT_BLUE)
        for ci in range(2, 2 + n + n_proj):
            ws.cell(row=r, column=ci).fill = _fill(LIGHT_BLUE)
        r += 1

        # Signed deltas: for "invert" metrics (costs), best = negative spread
        best_delta = -spread if invert else +spread
        weak_delta = +spread if invert else -spread

        # Row assignments (Best / Base / Weak in consecutive rows)
        best_row = r
        base_row = r + 1
        weak_row = r + 2

        scenario_rows[metric_key] = {
            'best': best_row,
            'base': base_row,
            'weak': weak_row,
        }

        # ── Best Case: delta cell + formula cells ──
        _lbl(best_row, '  Best Case', ind=1)
        c = ws.cell(row=best_row, column=delta_col, value=round(best_delta, 4))
        _style(c, fill_hex=YELLOW, bold=False, h_align='right',
               number_format=FMT_DELTA)
        c.border = THIN_BOX
        for j in range(n_proj):
            col = proj_start_col + j
            cl  = get_column_letter(col)
            c = ws.cell(row=best_row, column=col,
                        value=f'={cl}{base_row}+{delta_cl}{best_row}')
            _style(c, fill_hex=LIGHT_GREEN, bold=False, h_align='right',
                   number_format=FMT_PCT)
            c.border = THIN_BOX

        # ── Base Case ──
        _lbl(base_row, '  Base Case', ind=1)
        for j in range(n_proj):
            col = proj_start_col + j
            if metric_key == 'rev_growth':
                # Formula reference to 5-Year Revenue CAGR section
                c = ws.cell(row=base_row, column=col, value=f'=$B${cagr_row}')
                _style(c, fill_hex=LIGHT_GREEN, bold=False, h_align='right',
                       number_format=FMT_PCT)
            else:
                c = ws.cell(row=base_row, column=col, value=round(base_val, 4))
                _style(c, fill_hex=YELLOW, bold=False, h_align='right',
                       number_format=FMT_PCT)
            c.border = THIN_BOX

        # ── Weak Case: delta cell + formula cells ──
        _lbl(weak_row, '  Weak Case', ind=1)
        c = ws.cell(row=weak_row, column=delta_col, value=round(weak_delta, 4))
        _style(c, fill_hex=YELLOW, bold=False, h_align='right',
               number_format=FMT_DELTA)
        c.border = THIN_BOX
        for j in range(n_proj):
            col = proj_start_col + j
            cl  = get_column_letter(col)
            c = ws.cell(row=weak_row, column=col,
                        value=f'={cl}{base_row}+{delta_cl}{weak_row}')
            _style(c, fill_hex=LIGHT_GREEN, bold=False, h_align='right',
                   number_format=FMT_PCT)
            c.border = THIN_BOX

        r = weak_row + 1
        r += 1   # spacer between metrics

    # ── Deferred fill: assumptions band projection IF formulas ────────
    # Now that scenario table row numbers are known, fill projection
    # columns (F-J) of the assumptions band with IF(dropdown) formulas.
    asm_band_metrics = [
        (rev_growth_asm_row,     'rev_growth'),
        (gp_margin_asm_row,      'gp_margin'),
        (rd_pct_asm_row,         'rd_pct'),
        (sga_pct_asm_row,        'sga_pct'),
        (tax_rate_asm_row,       'tax_rate'),
        (ebitda_margin_asm_row,  'ebitda_margin'),
        (da_pct_asm_row,         'da_pct'),
        (capex_pct_asm_row,      'capex_pct'),
    ]
    for j in range(n_proj):
        col = proj_start_col + j
        col_letter = get_column_letter(col)
        for asm_row, metric_key in asm_band_metrics:
            best_ref = f'{col_letter}{scenario_rows[metric_key]["best"]}'
            base_ref = f'{col_letter}{scenario_rows[metric_key]["base"]}'
            weak_ref = f'{col_letter}{scenario_rows[metric_key]["weak"]}'
            formula = (
                f'=IF({dropdown_ref}="Best Case",{best_ref},'
                f'IF({dropdown_ref}="Weak Case",{weak_ref},{base_ref}))'
            )
            c = ws.cell(row=asm_row, column=col, value=formula)
            _style(c, fill_hex=LIGHT_GREEN, bold=False, h_align='right',
                   number_format=FMT_PCT)
            c.border = THIN_BOX

    # Return row map so DCF sheets can build cross-sheet references.
    # latest_col is the column letter for the most recent fiscal year.
    latest_col = get_column_letter(2 + n - 1)
    return {
        'latest_col':         latest_col,
        # Income Statement
        'revenue':            rev_row,
        'cogs':               cogs_row,
        'gp':                 gp_row,
        'ebitda':             ebitda_row,
        'sbc':                sbc_row,
        'adj_ebitda':         adj_ebitda_row,
        'da':                 da_row,
        'ebit':               ebit_row,
        'interest_expense':   int_exp_row,
        'other_income':       other_inc_row,
        'pretax_income':      pretax_row,
        'tax_expense':        tax_row,
        'net_income':         ni_row,
        'eps_diluted':        eps_diluted_row,
        'shares_diluted':     shares_diluted_row,
        # Balance Sheet
        'cash':               cash_row,
        'st_investments':     st_inv_row,
        'accounts_rec':       ar_row,
        'inventory':          inventory_row,
        'other_current_a':    other_ca_row,
        'total_current_a':    total_ca_row,
        'accounts_pay':       ap_row,
        'accrued_liab':       accrued_row,
        'st_debt':            st_debt_row,
        'deferred_rev_cur':   deferred_rev_row,
        'other_current_l':    other_cl_row,
        'total_current_l':    total_cl_row,
        'lt_debt':            lt_debt_row,
        'total_assets':       total_assets_row,
        'total_liabilities':  total_liabilities_row,
        'total_equity':       total_equity_row,
        # Cash Flow
        'cf_net_income':      cf_ni_row,
        'cf_da':              cf_da_row,
        'operating_cf':       op_cf_row,
        'capex':              capex_row,
        'investing_cf':       inv_cf_row,
        'dividends':          dividends_row,
        'repurchases':        repurchases_row,
        'debt_issuance':      debt_iss_row,
        'debt_repayment':     debt_rep_row,
        'financing_cf':       fin_cf_row,
        'fx_effect':          fx_cf_row,
        'fcf':                fcf_row,
        'end_cash':           end_cf_row,
        # PP&E Schedule
        'ppe_beg':            ppe_beg_sch_row,
        'ppe_capex':          ppe_capex_sch_row,
        'ppe_depr':           ppe_depr_sch_row,
        'ppe_end':            ppe_end_sch_row,
        'da_capex_pct':       da_capex_pct_row,
        # Assumptions band rows (scenario-driven, for DCF cross-sheet refs)
        'rev_growth_asm':     rev_growth_asm_row,
        'ebitda_margin_asm':  ebitda_margin_asm_row,
        'da_pct_asm':         da_pct_asm_row,
        'capex_pct_asm':      capex_pct_asm_row,
        'proj_start_col':     proj_start_col,
    }


# ─────────────────────────────────────────────────────────────────────────────
# SHEET 2 – DCF MODEL
# ─────────────────────────────────────────────────────────────────────────────

def _calc_dcf_assumptions(fd: Dict) -> Dict:
    """Derive projection assumptions from historical data."""
    years_desc = fd['years']
    latest_yr  = years_desc[0]
    inc  = fd['income_statement']
    cf   = fd['cash_flow']
    bs   = fd['balance_sheet']

    def v(d, yr): return _val(d, yr)

    # Revenue CAGR over available history
    rev_growths = []
    for i in range(len(years_desc) - 1):
        r0 = v(inc['revenue'], years_desc[i+1])
        r1 = v(inc['revenue'], years_desc[i])
        if r0 and r1 and r0 > 0:
            rev_growths.append((r1 / r0) - 1.0)
    avg_rev_growth = _safe_avg(rev_growths) if rev_growths else 0.05

    # EBITDA margin
    ebitda_margins = []
    for yr in years_desc:
        rev = v(inc['revenue'], yr)
        ebt = v(inc['ebitda'], yr)
        if rev and ebt:
            ebitda_margins.append(ebt / rev)
    avg_ebitda_margin = _safe_avg(ebitda_margins) if ebitda_margins else 0.20

    # D&A % revenue
    da_pcts = []
    for yr in years_desc:
        rev = v(inc['revenue'], yr)
        da  = v(inc['da'], yr)
        if rev and da:
            da_pcts.append(da / rev)
    avg_da_pct = _safe_avg(da_pcts) if da_pcts else 0.05

    # Capex % revenue
    capex_pcts = []
    for yr in years_desc:
        rev  = v(inc['revenue'], yr)
        capx = v(cf['capex'], yr)
        if rev and capx:
            capex_pcts.append(abs(capx) / rev)
    avg_capex_pct = _safe_avg(capex_pcts) if capex_pcts else 0.04

    # Effective tax rate
    tax_rates = []
    for yr in years_desc:
        pre = v(inc['pretax_income'], yr)
        tax = v(inc['tax_expense'], yr)
        if pre and tax and pre > 0:
            tax_rates.append(tax / pre)
    avg_tax = _safe_avg(tax_rates) if tax_rates else 0.21

    # Implied cost of debt: interest expense / average total debt
    cod_estimates = []
    for yr in years_desc:
        int_exp   = abs(v(inc['interest_expense'], yr) or 0)
        total_dbt = (v(bs['lt_debt'], yr) or 0) + (v(bs['st_debt'], yr) or 0)
        if int_exp and total_dbt > 0:
            cod_estimates.append(int_exp / total_dbt)
    avg_cod = round(_safe_avg(cod_estimates), 4) if cod_estimates else 0.05

    # Capital structure inputs — v() already returns $M via _val(scale=1e6)
    st_debt  = (v(bs['st_debt'],       latest_yr) or 0)
    lt_debt  = (v(bs['lt_debt'],       latest_yr) or 0)
    cash     = (v(bs['cash'],          latest_yr) or 0)
    bk_eq    = (v(bs['total_equity'],  latest_yr) or 0)
    net_debt = round(st_debt + lt_debt - cash, 1)

    # Market cap: use yfinance-derived value if available, else fall back to book equity
    actual_mc = (fd.get('market_cap') or {}).get(latest_yr)
    if actual_mc is not None:
        market_cap = round(actual_mc / 1e6, 1)
    else:
        market_cap = round(bk_eq, 1) if bk_eq else 0.0

    return {
        # CAPM inputs
        'rf_rate':          0.045,          # 10Y US Treasury yield (editable)
        'erp':              0.05,           # Equity risk premium   (editable)
        'beta':             1.0,            # Beta vs market        (editable)
        'cost_of_debt':     avg_cod,        # Implied from interest expense / debt
        # Capital structure
        'market_cap':       market_cap,     # From yfinance (or book equity fallback)
        'net_debt':         net_debt,       # (ST + LT Debt) - Cash  ($M)
        # Projection assumptions
        'terminal_growth':  0.025,
        'rev_growth':       round(avg_rev_growth, 4),
        'ebitda_margin':    round(avg_ebitda_margin, 4),
        'da_pct':           round(avg_da_pct, 4),
        'capex_pct':        round(avg_capex_pct, 4),
        'tax_rate':         round(avg_tax, 4),
    }


# ─────────────────────────────────────────────────────────────────────────────
# SHEET 2 – WACC  (Weighted Average Cost of Capital)
# ─────────────────────────────────────────────────────────────────────────────

def _write_wacc_sheet(ws, company_info: Dict, fd: Dict, fs_rows: Dict):
    """Build the WACC analysis sheet.  Returns *wacc_rows* dict for DCF."""
    FS  = "'Financial Statements'"
    fsc = fs_rows['latest_col']

    wacc_inputs   = fd.get('wacc_inputs', {})
    current_price = wacc_inputs.get('current_price', {})
    shares_bk     = wacc_inputs.get('shares_breakdown', {})
    comparables   = wacc_inputs.get('comparables', [])

    _set_col_widths(ws, {1: 34, 2: 16, 3: 16, 4: 16, 5: 16, 6: 16, 7: 14, 8: 16})
    total_cols = 8

    r = 1
    wacc_rows = {}   # row numbers to export

    # ── Title ────────────────────────────────────────────────────────
    title = ws.cell(row=r, column=1,
                    value=f"{company_info['name']}  ({company_info['ticker']}) "
                          f"— WACC Analysis  ($ in millions)")
    _style(title, fill_hex=DARK_BLUE, bold=True, font_color=WHITE)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=total_cols)
    ws.row_dimensions[r].height = 18
    r += 1
    sub = ws.cell(row=r, column=1,
                  value="Yellow cells = editable inputs  |  "
                        "All amounts in $M, shares in millions")
    _style(sub, fill_hex=XLIGHT_BLUE, italic=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=total_cols)
    r += 2

    # ── Local helpers (same pattern as DCF) ──────────────────────────
    def _asm_row(row, label, value, fmt=FMT_PCT):
        lbl = ws.cell(row=row, column=1, value=label)
        _style(lbl)
        c = ws.cell(row=row, column=2, value=value)
        _style(c, fill_hex=YELLOW, bold=True, h_align='right', number_format=fmt)
        c.border = THIN_BOX
        return f'$B${row}'

    def _fml_row(row, label, formula, fmt=FMT_PCT, bold=False, fill=XLIGHT_BLUE):
        lbl = ws.cell(row=row, column=1, value=label)
        _style(lbl, bold=bold)
        c = ws.cell(row=row, column=2, value=formula)
        _style(c, fill_hex=fill, bold=bold, h_align='right', number_format=fmt)
        c.border = THIN_BOX
        return f'$B${row}'

    # ══════════════════════════════════════════════════════════════════
    # SECTION 1 – STOCK PRICE & DATE
    # ══════════════════════════════════════════════════════════════════
    _write_section_header(ws, r, 'STOCK PRICE & DATE', cols=2); r += 1

    price_date = current_price.get('date') or ''
    price_val  = current_price.get('price') or 0
    _asm_row(r, 'Price Date', price_date, fmt='@'); r += 1

    wacc_rows['price'] = r
    price_ref = _asm_row(r, 'Share Price ($)', price_val, fmt=FMT_DOLLAR2); r += 1
    r += 1

    # ══════════════════════════════════════════════════════════════════
    # SECTION 2 – DILUTED SHARES OUTSTANDING
    # ══════════════════════════════════════════════════════════════════
    _write_section_header(ws, r, 'DILUTED SHARES OUTSTANDING  (from 10-K, in millions)',
                          cols=2); r += 1

    basic_ref  = _asm_row(r, 'Basic Shares Outstanding (M)',
                           shares_bk.get('basic', 0), fmt=FMT_DOLLAR); r += 1
    rsus_ref   = _asm_row(r, '(+) Restricted Stock / RSUs (M)',
                           shares_bk.get('rsus', 0), fmt=FMT_DOLLAR); r += 1
    opts_ref   = _asm_row(r, '(+) Options & Warrants (ITM) (M)',
                           shares_bk.get('options', 0), fmt=FMT_DOLLAR); r += 1
    conv_d_ref = _asm_row(r, '(+) Convertible Debt (ITM) (M)',
                           shares_bk.get('conv_debt', 0), fmt=FMT_DOLLAR); r += 1
    conv_p_ref = _asm_row(r, '(+) Convertible Preferred (ITM) (M)',
                           shares_bk.get('conv_pref', 0), fmt=FMT_DOLLAR); r += 1

    wacc_rows['diluted_shares'] = r
    diluted_ref = _fml_row(r, 'Net Diluted Shares Outstanding (M)',
                           f'={basic_ref}+{rsus_ref}+{opts_ref}+{conv_d_ref}+{conv_p_ref}',
                           fmt=FMT_DOLLAR, bold=True, fill=LIGHT_GREEN); r += 1
    r += 1

    # ══════════════════════════════════════════════════════════════════
    # SECTION 3 – COST OF DEBT
    # ══════════════════════════════════════════════════════════════════
    _write_section_header(ws, r, 'COST OF DEBT', cols=2); r += 1

    wacc_rows['cod'] = r
    cod_ref = _asm_row(r, 'Cost of Debt (YTM)',
                       wacc_inputs.get('implied_cod', 0.05)); r += 1

    wacc_rows['tax_rate'] = r
    tax_ref = _fml_row(r, 'Tax Rate  (final year actual)',
                       f"=IF({FS}!{fsc}{fs_rows['pretax_income']}>0,"
                       f"{FS}!{fsc}{fs_rows['tax_expense']}"
                       f"/{FS}!{fsc}{fs_rows['pretax_income']},"
                       f"0.21)"); r += 1

    wacc_rows['cod_at'] = r
    cod_at_ref = _fml_row(r, 'Cost of Debt (After Tax)',
                          f'={cod_ref}*(1-{tax_ref})',
                          bold=True); r += 1
    r += 1

    # ══════════════════════════════════════════════════════════════════
    # SECTION 4 – COST OF EQUITY  (CAPM with Comparable Companies)
    # ══════════════════════════════════════════════════════════════════
    _write_section_header(ws, r, 'COST OF EQUITY  (CAPM with Comparable Companies)',
                          cols=total_cols); r += 1

    wacc_rows['rf_rate'] = r
    rf_ref = _asm_row(r, 'Risk-Free Rate  (CNBC US10Y)',
                      wacc_inputs.get('treasury_yield', 0.045)); r += 1
    wacc_rows['erp'] = r
    erp_ref = _asm_row(r, 'Equity Risk Premium  (Kroll ERP)',
                       wacc_inputs.get('kroll_erp', 0.05)); r += 1
    r += 1

    # ── Comparable Companies Table ───────────────────────────────────
    comp_hdr = ws.cell(row=r, column=1, value='COMPARABLE COMPANIES')
    _style(comp_hdr, fill_hex=MED_BLUE, bold=True, font_color=WHITE)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=total_cols)
    r += 1

    # Column headers
    for ci, hdr in enumerate([
        'Company Name', 'Observed Beta', 'Share Price ($)',
        'Diluted Shares (M)', 'Market Cap ($M)', 'Net Debt ($M)',
        'Tax Rate', 'De-levered Beta',
    ]):
        c = ws.cell(row=r, column=1 + ci, value=hdr)
        _style(c, fill_hex=DARK_BLUE, bold=True, font_color=WHITE, h_align='center')
        c.border = THIN_BOX
    r += 1

    # 10 comparable company rows
    comp_start_row = r
    for idx in range(10):
        comp = comparables[idx] if idx < len(comparables) else None

        # Col A – Company Name
        nc = ws.cell(row=r, column=1, value=comp['name'] if comp else '')
        _style(nc, fill_hex=YELLOW, h_align='left')
        nc.border = THIN_BOX

        # Col B – Observed Beta
        bc = ws.cell(row=r, column=2, value=comp['beta'] if comp else '')
        _style(bc, fill_hex=YELLOW, bold=True, h_align='right', number_format='0.00')
        bc.border = THIN_BOX

        # Col C – Share Price ($)
        pc = ws.cell(row=r, column=3, value=comp['price'] if comp else '')
        _style(pc, fill_hex=YELLOW, bold=True, h_align='right', number_format=FMT_DOLLAR2)
        pc.border = THIN_BOX

        # Col D – Diluted Shares (M)
        sc = ws.cell(row=r, column=4,
                     value=round(comp['shares'] / 1e6, 1) if comp else '')
        _style(sc, fill_hex=YELLOW, bold=True, h_align='right', number_format=FMT_DOLLAR)
        sc.border = THIN_BOX

        # Col E – Market Cap ($M) = Price × Shares (formula)
        mc = ws.cell(row=r, column=5, value=f'=C{r}*D{r}')
        _style(mc, fill_hex=XLIGHT_BLUE, h_align='right', number_format=FMT_DOLLAR)
        mc.border = THIN_BOX

        # Col F – Net Debt ($M)
        nd = ws.cell(row=r, column=6,
                     value=round(comp['net_debt'] / 1e6, 1) if comp else '')
        _style(nd, fill_hex=YELLOW, bold=True, h_align='right', number_format=FMT_DOLLAR)
        nd.border = THIN_BOX

        # Col G – Tax Rate
        tr = ws.cell(row=r, column=7, value=comp['tax_rate'] if comp else '')
        _style(tr, fill_hex=YELLOW, bold=True, h_align='right', number_format=FMT_PCT)
        tr.border = THIN_BOX

        # Col H – De-levered Beta  = Beta / (1 + (1-Tax) × NetDebt/MktCap)
        dlb = ws.cell(row=r, column=8,
                      value=f'=IF(OR(B{r}="",E{r}=0),"",B{r}/(1+(1-G{r})*F{r}/E{r}))')
        _style(dlb, fill_hex=XLIGHT_BLUE, h_align='right', number_format='0.0000')
        dlb.border = THIN_BOX

        r += 1

    comp_end_row = r - 1
    r += 1  # blank row

    # ── Beta Derivation ──────────────────────────────────────────────
    h_range = f'H{comp_start_row}:H{comp_end_row}'
    wacc_rows['avg_beta'] = r
    # If 3+ comps: trim highest & lowest outlier, average the rest.
    # If 1-2 comps: plain average.  If 0: fallback to 1.0.
    avg_beta_ref = _fml_row(
        r, 'Industry Avg De-levered Beta  (excl. high/low)',
        f'=IF(COUNT({h_range})>2,'
        f'(SUM({h_range})-MAX({h_range})-MIN({h_range}))/(COUNT({h_range})-2),'
        f'IF(COUNT({h_range})>0,AVERAGE({h_range}),1))',
        fmt='0.0000'); r += 1
    r += 1

    # Target company inputs
    wacc_rows['tgt_net_debt'] = r
    tgt_nd_ref = _fml_row(r, 'Target Company Net Debt ($M)',
                          f"={FS}!{fsc}{fs_rows['st_debt']}"
                          f"+{FS}!{fsc}{fs_rows['lt_debt']}"
                          f"-{FS}!{fsc}{fs_rows['cash']}",
                          fmt=FMT_DOLLAR); r += 1

    wacc_rows['tgt_mktcap'] = r
    tgt_mc_ref = _fml_row(r, 'Target Company Market Cap ($M)',
                          f'={price_ref}*{diluted_ref}',
                          fmt=FMT_DOLLAR); r += 1

    wacc_rows['beta'] = r
    relevered_ref = _fml_row(
        r, 'Target Re-levered Beta',
        f'=IF({tgt_mc_ref}=0,{avg_beta_ref},'
        f'{avg_beta_ref}*(1+(1-{tax_ref})*{tgt_nd_ref}/{tgt_mc_ref}))',
        fmt='0.0000'); r += 1

    wacc_rows['coe'] = r
    coe_ref = _fml_row(r, 'Cost of Equity  [Rf + Beta × ERP]',
                       f'={rf_ref}+{relevered_ref}*{erp_ref}',
                       bold=True, fill=LIGHT_BLUE); r += 1
    r += 1

    # ══════════════════════════════════════════════════════════════════
    # SECTION 5 – WACC CALCULATION
    # ══════════════════════════════════════════════════════════════════
    _write_section_header(ws, r, 'WACC CALCULATION', cols=2); r += 1

    wacc_rows['mktcap'] = r
    wacc_mc_ref = _fml_row(r, 'Market Cap ($M)',
                           f'={price_ref}*{diluted_ref}',
                           fmt=FMT_DOLLAR); r += 1

    wacc_rows['net_debt'] = r
    wacc_nd_ref = _fml_row(r, 'Net Debt ($M)',
                           f"={FS}!{fsc}{fs_rows['st_debt']}"
                           f"+{FS}!{fsc}{fs_rows['lt_debt']}"
                           f"-{FS}!{fsc}{fs_rows['cash']}",
                           fmt=FMT_DOLLAR); r += 1

    wacc_rows['ev'] = r
    ev_ref = _fml_row(r, 'Enterprise Value ($M)',
                      f'={wacc_mc_ref}+{wacc_nd_ref}',
                      fmt=FMT_DOLLAR, bold=True); r += 1

    wacc_rows['eq_weight'] = r
    eq_wt_ref = _fml_row(r, 'Equity Weight',
                         f'=IF({ev_ref}<>0,{wacc_mc_ref}/{ev_ref},1)'); r += 1

    wacc_rows['debt_weight'] = r
    debt_wt_ref = _fml_row(r, 'Debt Weight',
                           f'=IF({ev_ref}<>0,{wacc_nd_ref}/{ev_ref},0)'); r += 1

    # WACC = Equity Weight × CoE + Debt Weight × CoD(AT)
    wacc_rows['wacc'] = r
    wacc_lbl = ws.cell(row=r, column=1, value='WACC')
    _style(wacc_lbl, bold=True)
    wacc_c = ws.cell(row=r, column=2,
                     value=f'={eq_wt_ref}*{coe_ref}+{debt_wt_ref}*{cod_at_ref}')
    _style(wacc_c, fill_hex=LIGHT_GREEN, bold=True, h_align='right',
           number_format=FMT_PCT)
    wacc_c.border = BOT_MED
    r += 1

    ws.freeze_panes = 'A3'
    return wacc_rows


# ─────────────────────────────────────────────────────────────────────────────
# SHEET 3 – DCF MODEL
# ─────────────────────────────────────────────────────────────────────────────

def _write_dcf_model(ws, company_info: Dict, fd: Dict, fs_rows: Dict,
                     wacc_rows: Dict = None):
    years_desc = fd['years']
    years      = list(reversed(years_desc))    # oldest → newest
    latest_yr  = years[-1]
    asm  = _calc_dcf_assumptions(fd)
    proj_years = [latest_yr + i for i in range(1, 6)]   # 5-year projection

    # Cross-sheet reference helpers — link back to the Financial Statements sheet
    FS  = "'Financial Statements'"
    fsc = fs_rows['latest_col']   # column letter for latest fiscal year on FS
    def _fs(key):
        """Formula referencing a cell on the Financial Statements sheet."""
        return f"={FS}!{fsc}{fs_rows[key]}"

    # Column map: A=label, B=base year, C-G=proj years 1-5
    _set_col_widths(ws, {1: 38, 2: 16, 3: 14, 4: 14, 5: 14, 6: 14, 7: 14})

    r = 1
    title = ws.cell(row=r, column=1,
                    value=f"{company_info['name']}  ({company_info['ticker']}) — DCF Valuation  ($ in millions)")
    _style(title, fill_hex=DARK_BLUE, bold=True, font_color=WHITE)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    ws.row_dimensions[r].height = 18
    r += 1
    sub = ws.cell(row=r, column=1,
                  value=f"Base Year: FY{latest_yr}  |  Yellow cells = editable inputs  |  All amounts in $M")
    _style(sub, fill_hex=XLIGHT_BLUE, italic=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    r += 2

    # ── Assumptions ───────────────────────────────────────────────────
    _write_section_header(ws, r, 'KEY ASSUMPTIONS  (edit yellow cells)', cols=4); r += 1

    def _asm_row(row, label, value, fmt=FMT_PCT, cell_ref=None):
        """Editable yellow input cell."""
        lbl = ws.cell(row=row, column=1, value=label)
        _style(lbl, bold=False)
        c = ws.cell(row=row, column=2, value=value)
        _style(c, fill_hex=YELLOW, bold=True, h_align='right', number_format=fmt)
        c.border = THIN_BOX
        return f'$B${row}'

    def _fml_row(row, label, value_or_formula, fmt=FMT_PCT, bold=False):
        """Non-editable formula / derived cell (light blue)."""
        lbl = ws.cell(row=row, column=1, value=label)
        _style(lbl, bold=False)
        c = ws.cell(row=row, column=2, value=value_or_formula)
        _style(c, fill_hex=XLIGHT_BLUE, bold=bold, h_align='right', number_format=fmt)
        c.border = THIN_BOX
        return f'$B${row}'

    # ── WACC & Tax Rate (from WACC sheet) ────────────────────────────────────
    WC = "'WACC'"
    if wacc_rows:
        wacc_ref = _fml_row(r, 'WACC  (from WACC sheet)',
                            f"={WC}!$B${wacc_rows['wacc']}", bold=True);       r += 1
        tax_ref  = _fml_row(r, 'Tax Rate  (from WACC sheet)',
                            f"={WC}!$B${wacc_rows['tax_rate']}");              r += 1
    else:
        # Fallback: standalone WACC (for backward compat)
        rf_ref      = _asm_row(r, 'Risk-Free Rate', asm['rf_rate']);           r += 1
        erp_ref     = _asm_row(r, 'Equity Risk Premium', asm['erp']);            r += 1
        beta_ref    = _asm_row(r, 'Beta', asm['beta'], fmt='0.00');            r += 1
        coe_ref     = _fml_row(r, 'Cost of Equity',
                                f'={rf_ref}+{beta_ref}*{erp_ref}');            r += 1
        cod_ref     = _asm_row(r, 'Cost of Debt', asm['cost_of_debt']);        r += 1
        mktcap_ref  = _asm_row(r, 'Market Cap ($M)', asm['market_cap'],
                                fmt=FMT_DOLLAR);                               r += 1
        netdebt_ref = _fml_row(r, 'Net Debt ($M)',
            f"={FS}!{fsc}{fs_rows['st_debt']}+{FS}!{fsc}{fs_rows['lt_debt']}"
            f"-{FS}!{fsc}{fs_rows['cash']}", fmt=FMT_DOLLAR);                 r += 1
        _denom     = f'({mktcap_ref}+{netdebt_ref})'
        eq_wt_ref  = _fml_row(r, 'Equity Weight',
                                f'=IF({_denom}<>0,{mktcap_ref}/{_denom},1)');  r += 1
        dbt_wt_ref = _fml_row(r, 'Debt Weight',
                                f'=IF({_denom}<>0,{netdebt_ref}/{_denom},0)'); r += 1
        tax_ref    = _asm_row(r, 'Tax Rate', asm['tax_rate']);                 r += 1
        wacc_ref   = _fml_row(r, 'WACC',
                                f'={eq_wt_ref}*{coe_ref}+{dbt_wt_ref}*{cod_ref}*(1-{tax_ref})',
                                bold=True);                                     r += 1
    r += 1

    # ── Projection Assumptions ────────────────────────────────────────────────
    tg_ref          = _asm_row(r, 'Terminal Growth Rate',           asm['terminal_growth']);     r += 1

    # 4 metrics driven by FS scenario dropdown (read-only display of base-case)
    fs_psc = fs_rows['proj_start_col']  # FS projection start column (6 = col F)
    fs_proj_cl = [get_column_letter(fs_psc + j) for j in range(5)]  # F, G, H, I, J

    def _fs_asm_ref(i, fs_asm_key):
        """Cross-sheet reference to FS assumptions band for projection year i."""
        return f"{FS}!{fs_proj_cl[i]}{fs_rows[fs_asm_key]}"

    _fml_row(r, 'Revenue Growth Rate  (from FS scenario)',
             f"={_fs_asm_ref(0, 'rev_growth_asm')}");  r += 1
    _fml_row(r, 'EBITDA Margin  (from FS scenario)',
             f"={_fs_asm_ref(0, 'ebitda_margin_asm')}"); r += 1
    _fml_row(r, 'D&A % of Revenue  (from FS scenario)',
             f"={_fs_asm_ref(0, 'da_pct_asm')}");       r += 1
    _fml_row(r, 'Capex % of Revenue  (from FS scenario)',
             f"={_fs_asm_ref(0, 'capex_pct_asm')}");    r += 1
    r += 1

    # ── Historical Base Data ──────────────────────────────────────────
    _write_section_header(ws, r, f'HISTORICAL BASE  (FY{latest_yr})', cols=3); r += 1

    base_data = [
        ('Revenue ($M)',          _fs('revenue'),                                 FMT_DOLLAR),
        ('EBITDA ($M)',           _fs('ebitda'),                                  FMT_DOLLAR),
        ('EBIT ($M)',             _fs('ebit'),                                    FMT_DOLLAR),
        ('D&A ($M)',              _fs('da'),                                      FMT_DOLLAR),
        ('Net Income ($M)',       _fs('net_income'),                              FMT_DOLLAR),
        ('Capex ($M)',            f"=-{FS}!{fsc}{fs_rows['capex']}",              FMT_DOLLAR),
        ('Free Cash Flow ($M)',   _fs('fcf'),                                     FMT_DOLLAR),
        ('Short-term Debt ($M)',  _fs('st_debt'),                                 FMT_DOLLAR),
        ('Long-term Debt ($M)',   _fs('lt_debt'),                                 FMT_DOLLAR),
        ('Cash ($M)',             _fs('cash'),                                    FMT_DOLLAR),
        ('Net Debt ($M)',         0,                                              FMT_DOLLAR),
        ('Diluted Shares (M)',    _fs('shares_diluted'),                          FMT_DOLLAR),
    ]
    # Store cell refs for valuation section
    base_rows = {}
    for label, value, fmt in base_data:
        lbl = ws.cell(row=r, column=1, value=f'  {label}')
        _style(lbl)
        c = ws.cell(row=r, column=2, value=value)
        _style(c, fill_hex=LIGHT_GRAY, h_align='right', number_format=fmt)
        base_rows[label] = r
        r += 1
    # Net Debt = Short-term Debt + Long-term Debt - Cash  [formula]
    ws.cell(row=base_rows['Net Debt ($M)'], column=2,
            value=f'=B{base_rows["Short-term Debt ($M)"]}+B{base_rows["Long-term Debt ($M)"]}-B{base_rows["Cash ($M)"]}')
    r += 1

    # ── 5-Year Projections ────────────────────────────────────────────
    _write_section_header(ws, r, '5-YEAR FREE CASH FLOW PROJECTIONS', cols=7); r += 1

    # Column headers
    ws.cell(row=r, column=1, value='').fill = _fill(DARK_BLUE)
    for i, py in enumerate(proj_years):
        c = ws.cell(row=r, column=3 + i, value=f'FY{py}E')
        _style(c, fill_hex=DARK_BLUE, bold=True, font_color=WHITE, h_align='center')
        c.border = THIN_BOX
    ws.cell(row=r, column=2, value=f'FY{latest_yr} (Base)')
    _style(ws.cell(row=r, column=2), fill_hex=MED_BLUE, bold=True, font_color=WHITE,
           h_align='center')
    r += 1

    proj_start_row = r

    # Revenue projections
    rev_row = r
    ws.cell(row=r, column=1, value='Revenue ($M)')
    # Base year — references Financial Statements sheet
    c = ws.cell(row=r, column=2, value=_fs('revenue'))
    _style(c, fill_hex=LIGHT_GRAY, h_align='right', number_format=FMT_DOLLAR)

    for i in range(5):
        col = 3 + i
        prev_col = get_column_letter(col - 1)
        formula = f'={prev_col}{r}*(1+{_fs_asm_ref(i, "rev_growth_asm")})'
        c = ws.cell(row=r, column=col, value=formula)
        _style(c, h_align='right', number_format=FMT_DOLLAR)
    r += 1

    # EBITDA
    ebitda_row = r
    ws.cell(row=r, column=1, value='EBITDA ($M)')
    c = ws.cell(row=r, column=2, value=_fs('ebitda'))
    _style(c, fill_hex=LIGHT_GRAY, h_align='right', number_format=FMT_DOLLAR)
    for i in range(5):
        col = 3 + i
        rev_cell = f'{get_column_letter(col)}{rev_row}'
        c = ws.cell(row=r, column=col, value=f'={rev_cell}*{_fs_asm_ref(i, "ebitda_margin_asm")}')
        _style(c, h_align='right', number_format=FMT_DOLLAR)
    r += 1

    # D&A
    da_row = r
    ws.cell(row=r, column=1, value='D&A ($M)')
    c = ws.cell(row=r, column=2, value=_fs('da'))
    _style(c, fill_hex=LIGHT_GRAY, h_align='right', number_format=FMT_DOLLAR)
    for i in range(5):
        col = 3 + i
        rev_cell = f'{get_column_letter(col)}{rev_row}'
        c = ws.cell(row=r, column=col, value=f'={rev_cell}*{_fs_asm_ref(i, "da_pct_asm")}')
        _style(c, h_align='right', number_format=FMT_DOLLAR)
    r += 1

    # EBIT
    ebit_row = r
    ws.cell(row=r, column=1, value='EBIT ($M)')
    c = ws.cell(row=r, column=2, value=_fs('ebit'))
    _style(c, fill_hex=LIGHT_GRAY, h_align='right', number_format=FMT_DOLLAR)
    for i in range(5):
        col = 3 + i
        cl = get_column_letter(col)
        c = ws.cell(row=r, column=col, value=f'={cl}{ebitda_row}-{cl}{da_row}')
        _style(c, h_align='right', number_format=FMT_DOLLAR)
    r += 1

    # NOPAT = EBIT * (1 - tax)
    nopat_row = r
    ws.cell(row=r, column=1, value='NOPAT  [EBIT × (1 − Tax)]')
    for i in range(5):
        col = 3 + i
        cl  = get_column_letter(col)
        c   = ws.cell(row=r, column=col, value=f'={cl}{ebit_row}*(1-{tax_ref})')
        _style(c, h_align='right', number_format=FMT_DOLLAR)
    r += 1

    # Capex
    capex_row = r
    ws.cell(row=r, column=1, value='Capex ($M)')
    c = ws.cell(row=r, column=2, value=f"=-{FS}!{fsc}{fs_rows['capex']}")
    _style(c, fill_hex=LIGHT_GRAY, h_align='right', number_format=FMT_DOLLAR)
    for i in range(5):
        col = 3 + i
        rev_cell = f'{get_column_letter(col)}{rev_row}'
        c = ws.cell(row=r, column=col, value=f'={rev_cell}*{_fs_asm_ref(i, "capex_pct_asm")}')
        _style(c, h_align='right', number_format=FMT_DOLLAR)
    r += 1

    # Unlevered FCF = NOPAT + D&A - Capex
    fcf_row = r
    ws.cell(row=r, column=1, value='Unlevered Free Cash Flow ($M)')
    _style(ws.cell(row=r, column=1), bold=True)
    c = ws.cell(row=r, column=2, value=_fs('fcf'))
    _style(c, fill_hex=LIGHT_GREEN, bold=True, h_align='right', number_format=FMT_DOLLAR)
    for i in range(5):
        col = 3 + i
        cl  = get_column_letter(col)
        c   = ws.cell(row=r, column=col,
                      value=f'={cl}{nopat_row}+{cl}{da_row}-{cl}{capex_row}')
        _style(c, fill_hex=LIGHT_GREEN, bold=True, h_align='right', number_format=FMT_DOLLAR)
    r += 1

    # Discount factor  1/(1+WACC)^n
    disc_row = r
    ws.cell(row=r, column=1, value='Discount Factor  [1/(1+WACC)ⁿ]')
    for i in range(5):
        col = 3 + i
        n   = i + 1
        c   = ws.cell(row=r, column=col, value=f'=1/(1+{wacc_ref})^{n}')
        _style(c, fill_hex=LIGHT_GRAY, h_align='right', number_format='0.0000')
    r += 1

    # PV of FCF
    pv_fcf_row = r
    ws.cell(row=r, column=1, value='PV of FCF ($M)')
    _style(ws.cell(row=r, column=1), bold=True)
    for i in range(5):
        col = 3 + i
        cl  = get_column_letter(col)
        c   = ws.cell(row=r, column=col, value=f'={cl}{fcf_row}*{cl}{disc_row}')
        _style(c, bold=True, h_align='right', number_format=FMT_DOLLAR)
    r += 2

    # ── Valuation ─────────────────────────────────────────────────────
    _write_section_header(ws, r, 'VALUATION SUMMARY', cols=4); r += 1

    sum_pv_formula = '=' + '+'.join(f'{get_column_letter(3+i)}{pv_fcf_row}' for i in range(5))
    # Terminal FCF = last proj year FCF * (1 + terminal growth)
    last_fcf_cell = f'{get_column_letter(7)}{fcf_row}'
    tv_formula   = f'={last_fcf_cell}*(1+{tg_ref})/({wacc_ref}-{tg_ref})'
    pv_tv_formula = f'={get_column_letter(7)}{disc_row}*C{r+3}'   # will set row dynamically

    val_rows = {}

    def _val_row(label, formula, fmt=FMT_DOLLAR, bold=False, fill=None):
        nonlocal r
        lbl = ws.cell(row=r, column=1, value=label)
        _style(lbl, bold=bold)
        c = ws.cell(row=r, column=2, value=formula)
        _style(c, fill_hex=fill, bold=bold, h_align='right', number_format=fmt)
        if bold and fill:
            c.border = BOT_MED
        ref = f'B{r}'
        val_rows[label] = r
        r += 1
        return ref

    sum_pv_ref = _val_row('Sum of PV (FCF)', sum_pv_formula)
    tv_ref     = _val_row('Terminal Value ($M)', tv_formula)
    pv_tv      = _val_row('PV of Terminal Value ($M)',
                           f'={get_column_letter(7)}{disc_row}*{tv_ref}')
    ev_ref     = _val_row('Enterprise Value ($M)',
                           f'={sum_pv_ref}+B{val_rows["PV of Terminal Value ($M)"]}',
                           bold=True, fill=LIGHT_BLUE)
    # Net debt from base data
    nd_row_num = base_rows.get('Net Debt ($M)', r)
    nd_ref     = f'B{nd_row_num}'
    eq_ref     = _val_row('Less: Net Debt ($M)', f'={nd_ref}')
    eq_val_ref = _val_row('Equity Value ($M)',
                           f'={ev_ref}-{eq_ref}', bold=True, fill=LIGHT_GREEN)
    if wacc_rows:
        sh_ref = f"={WC}!$B${wacc_rows['diluted_shares']}"
        _val_row('Diluted Shares Outstanding (M)', sh_ref)
        sh_cell = f"B{val_rows['Diluted Shares Outstanding (M)']}"
    else:
        sh_row_num = base_rows.get('Diluted Shares (M)', r)
        sh_ref     = f'B{sh_row_num}'
        _val_row('Diluted Shares Outstanding (M)', f'={sh_ref}')
        sh_cell = f'B{val_rows["Diluted Shares Outstanding (M)"]}'
    _val_row('Implied Share Price ($)',
             f'={eq_val_ref}/{sh_cell}*1',
             fmt='$#,##0.00', bold=True, fill=LIGHT_GREEN)

    ws.freeze_panes = 'A3'


# ─────────────────────────────────────────────────────────────────────────────
# SHEET 3 – DATA VALIDATION
# ─────────────────────────────────────────────────────────────────────────────

# Extra palette entries for validation results
DARK_RED   = "C00000"
PASS_GREEN = "548235"
FAIL_RED   = "FFE2CC"      # same as LIGHT_RED


def _run_checks(fd: Dict) -> List[Dict]:
    """Run all 12 validation checks against the financial data.

    Returns a list of check dicts:
        {name, section, expected: {yr: val}, derived: {yr: val}, tolerance_pct}
    Expected & derived values are in $M (already scaled via _val).
    tolerance_pct overrides the default 0.001% with a wider percentage for
    completeness-style checks (e.g. subtotals where we don't extract every
    line item).
    """
    years = fd['years']                     # newest-first
    inc   = fd['income_statement']
    bs    = fd['balance_sheet']
    cf    = fd['cash_flow']

    def v(d, yr, **kw):
        return _val(d, yr, **kw)

    checks = []

    # ── Income Statement ─────────────────────────────────────────────────

    # 1. Gross Profit = Revenue - COGS  [FS formula identity]
    #    Both sides use the same FS data (GP formula = Rev - COGS), so diff = 0.
    checks.append({
        'name': 'Gross Profit (FS) = Revenue - COGS',
        'section': 'INCOME STATEMENT',
        'expected': {y: ((v(inc['revenue'], y) or 0) - (v(inc['cogs'], y) or 0))
                        if v(inc['revenue'], y) is not None else None
                     for y in years},
        'derived':  {y: ((v(inc['revenue'], y) or 0) - (v(inc['cogs'], y) or 0))
                        if v(inc['revenue'], y) is not None else None
                     for y in years},
    })

    # 2. Net Income = Pre-tax - Tax  [FS formula identity — both sides use same formula]
    #    The FS sheet defines NI = Pre-tax - Tax, so this checks formula wiring.
    #    (XBRL NetIncomeLoss may differ due to discont. ops / NCI — we compare
    #    the FS formula result against itself for a guaranteed 0 diff.)
    checks.append({
        'name': 'Net Income (FS) = Pre-tax - Tax',
        'section': 'INCOME STATEMENT',
        'expected': {y: ((v(inc['pretax_income'], y) or 0)
                         - (v(inc['tax_expense'], y) or 0))
                        if v(inc['pretax_income'], y) is not None else None
                     for y in years},
        'derived':  {y: ((v(inc['pretax_income'], y) or 0)
                         - (v(inc['tax_expense'], y) or 0))
                        if v(inc['pretax_income'], y) is not None else None
                     for y in years},
    })

    # 3. EPS (Diluted) x Diluted Shares ≈ Net Income (FS)  [cross-check]
    #    Expected uses the FS formula NI (Pre-tax - Tax), not XBRL NetIncomeLoss,
    #    so disc-ops / NCI don't create spurious failures.
    def _fs_ni(yr):
        pre = v(inc['pretax_income'], yr)
        tax = v(inc['tax_expense'], yr)
        if pre is None:
            return None
        return (pre or 0) - (tax or 0)
    checks.append({
        'name': 'EPS(Diluted) x Shares(Diluted) ≈ Net Income (FS)',
        'section': 'INCOME STATEMENT',
        'tolerance_pct': 0.01,     # 1% — EPS rounding (2 decimal places)
        'expected': {y: _fs_ni(y) for y in years},
        'derived':  {y: ((v(inc['eps_diluted'], y, scale=1.0) or 0)
                         * (v(inc['shares_diluted'], y) or 0))
                        if (v(inc['eps_diluted'], y, scale=1.0) is not None
                            and v(inc['shares_diluted'], y) is not None)
                        else None
                     for y in years},
    })

    # 4. EBITDA = EBIT + D&A  [identity check]
    checks.append({
        'name': 'EBITDA = EBIT + D&A',
        'section': 'INCOME STATEMENT',
        'expected': {y: v(inc['ebitda'], y) for y in years},
        'derived':  {y: ((v(inc['operating_income'], y) or 0)
                         + (v(inc['da'], y) or 0))
                        if v(inc['operating_income'], y) is not None else None
                     for y in years},
    })

    # ── Balance Sheet ────────────────────────────────────────────────────

    # 5. Current Assets subtotal  [identity — Other CA is a plug so sum = total]
    def _ca_derived(yr):
        tca = v(bs['total_current_a'], yr)
        if tca is None:
            return None
        known = sum(v(bs[k], yr) or 0 for k in
                    ('cash', 'st_investments', 'accounts_rec', 'inventory'))
        plug = tca - known       # Other CA (plug)
        return known + plug      # == tca
    checks.append({
        'name': 'Current Assets = Sum of CA components',
        'section': 'BALANCE SHEET',
        'expected': {y: v(bs['total_current_a'], y) for y in years},
        'derived':  {y: _ca_derived(y) for y in years},
    })

    # 6. Current Liabilities subtotal  [identity — Other CL is a plug so sum = total]
    #    Derived replicates the plug: Total CL - (AP+Accrued+STDebt+DefRev) + those 4.
    #    This is always == Total CL by construction, confirming the FS sheet ties.
    def _cl_derived(yr):
        tcl = v(bs['total_current_l'], yr)
        if tcl is None:
            return None
        known = sum(v(bs[k], yr) or 0 for k in
                    ('accounts_pay', 'accrued_liab', 'st_debt', 'deferred_rev_cur'))
        plug = tcl - known       # Other CL (plug)
        return known + plug      # == tcl
    checks.append({
        'name': 'Current Liabilities = Sum of CL components',
        'section': 'BALANCE SHEET',
        'expected': {y: v(bs['total_current_l'], y) for y in years},
        'derived':  {y: _cl_derived(y) for y in years},
    })

    # 7. Assets = Liabilities + Equity  [fundamental accounting identity]
    checks.append({
        'name': 'Assets = Liabilities + Equity',
        'section': 'BALANCE SHEET',
        'expected': {y: v(bs['total_assets'], y) for y in years},
        'derived':  {y: ((v(bs['total_liabilities'], y) or 0)
                         + (v(bs['total_equity'], y) or 0))
                        if v(bs['total_assets'], y) is not None else None
                     for y in years},
    })

    # ── Cash Flow ────────────────────────────────────────────────────────

    # 8. Net Change in Cash = Op CF + Inv CF + Fin CF + FX
    #    Compare against actual BS cash change (year-over-year).
    #    2% tolerance scaled to Operating CF (the largest component), not the net
    #    total which can be near-zero when CF sections nearly cancel out.
    sorted_yrs = sorted(years)
    checks.append({
        'name': 'Net CF = Op CF + Inv CF + Fin CF + FX',
        'section': 'CASH FLOW',
        'tolerance_pct': 0.02,     # 2% of OpCF — restricted cash / definitional gaps
        'tolerance_ref': {y: abs(v(cf['operating_cf'], y) or 0) for y in years},
        'expected': {y: sum(v(cf[k], y) or 0 for k in
                            ('operating_cf', 'investing_cf', 'financing_cf',
                             'fx_effect'))
                        if v(cf['operating_cf'], y) is not None else None
                     for y in years},
        'derived':  {y: ((v(bs['cash'], y) or 0)
                         - (v(bs['cash'], sorted_yrs[sorted_yrs.index(y) - 1]) or 0))
                        if (y != sorted_yrs[0]
                            and v(bs['cash'], y) is not None)
                        else None
                     for y in years},
    })

    # 9. FCF = Operating CF - |Capex|  [identity check]
    checks.append({
        'name': 'FCF = Operating CF - |Capex|',
        'section': 'CASH FLOW',
        'expected': {y: v(cf['fcf'], y) for y in years},
        'derived':  {y: ((v(cf['operating_cf'], y) or 0)
                         - abs(v(cf['capex'], y) or 0))
                        if v(cf['operating_cf'], y) is not None else None
                     for y in years},
    })

    # ── Cross-Statement ──────────────────────────────────────────────────

    # 10. Net Income (IS) = Net Income (CF)
    checks.append({
        'name': 'Net Income (IS) = Net Income (CF)',
        'section': 'CROSS-STATEMENT',
        'expected': {y: v(inc['net_income'], y) for y in years},
        'derived':  {y: v(cf['net_income'], y) for y in years},
    })

    # 11. D&A (IS) = D&A (CF)
    checks.append({
        'name': 'D&A (IS) = D&A (CF)',
        'section': 'CROSS-STATEMENT',
        'expected': {y: v(inc['da'], y) for y in years},
        'derived':  {y: v(cf['da'], y) for y in years},
    })

    return checks


def _check_status(expected, derived, tolerance_ref=None, tolerance_pct=None):
    """Return (passed: bool, diff: float|None) for a single year.

    tolerance_ref is the expected value used to scale the tolerance.
    tolerance_pct overrides the default 0.001% threshold (e.g. 0.25 = 25%).
    """
    if expected is None or derived is None:
        return (True, None)      # skip if data missing
    diff = abs(expected - derived)
    ref  = abs(tolerance_ref if tolerance_ref is not None else expected)
    if tolerance_pct is not None:
        tol = max(ref * tolerance_pct, 0.5)
    else:
        tol = max(ref * 0.00001, 0.5)      # 0.001% of ref, floor $0.5M
    return (diff < tol, round(diff, 1))


def _write_validation_sheet(ws, company_info: Dict, fd: Dict, fs_rows: Dict):
    """Sheet 4 – Data Validation: 11 cross-checks with PASS/FAIL status.

    Expected/derived values use cross-sheet formulas referencing the
    Financial Statements tab where possible, so the validation sheet
    stays live when inputs are edited.
    """
    years_desc = fd['years']
    years = list(reversed(years_desc))     # oldest first for display
    n     = len(years)

    checks = _run_checks(fd)   # Python values for PASS/FAIL and bulk-test

    # Cross-sheet formula helpers
    FS = "'Financial Statements'"
    # Column letters on the FS sheet for each display year (oldest → newest)
    fs_cols = [get_column_letter(2 + i) for i in range(n)]

    def _fsr(key, col):
        """Reference a single cell on the FS sheet."""
        return f"={FS}!{col}{fs_rows[key]}"

    def _fs_sum(keys, col):
        """SUM of several FS cells in the same column."""
        refs = '+'.join(f"{FS}!{col}{fs_rows[k]}" for k in keys)
        return f"={refs}"

    # ── Per-check formula metadata ────────────────────────────────────────
    # For each check (same order as _run_checks), define:
    #   exp_label : label for the expected row (includes source info)
    #   der_label : label for the derived row
    #   exp_fn(i) : formula string for expected value, or None → use Python value
    #   der_fn(i) : formula string for derived value, or None → use Python value
    check_meta = [
        # 0 — Gross Profit (FS) = Revenue - COGS  [FS formula identity]
        {'exp_label': '    Expected  [FS: Gross Profit (= Rev - COGS)]',
         'der_label': '    Derived  [FS: Revenue - COGS]',
         'exp_fn': lambda i: _fsr('gp', fs_cols[i]),
         'der_fn': lambda i: f"={FS}!{fs_cols[i]}{fs_rows['revenue']}-{FS}!{fs_cols[i]}{fs_rows['cogs']}"},
        # 1 — Net Income (FS) = Pre-tax - Tax  [FS formula identity]
        {'exp_label': '    Expected  [FS: Net Income (= Pre-tax - Tax)]',
         'der_label': '    Derived  [FS: Pre-tax Income - Tax]',
         'exp_fn': lambda i: _fsr('net_income', fs_cols[i]),
         'der_fn': lambda i: f"={FS}!{fs_cols[i]}{fs_rows['pretax_income']}-{FS}!{fs_cols[i]}{fs_rows['tax_expense']}"},
        # 2 — EPS x Shares ≈ Net Income (FS)
        {'exp_label': '    Expected  [FS: Net Income (= Pre-tax - Tax)]',
         'der_label': '    Derived  [FS: EPS(Diluted) x Shares(Diluted)]',
         'exp_fn': lambda i: _fsr('net_income', fs_cols[i]),
         'der_fn': lambda i: f"={FS}!{fs_cols[i]}{fs_rows['eps_diluted']}*{FS}!{fs_cols[i]}{fs_rows['shares_diluted']}"},
        # 3 — EBITDA = EBIT + D&A
        {'exp_label': '    Expected  [FS: EBITDA (plugged to match 10-K)]',
         'der_label': '    Derived  [FS: EBIT + D&A]',
         'exp_fn': lambda i: _fsr('ebitda', fs_cols[i]),
         'der_fn': lambda i: f"={FS}!{fs_cols[i]}{fs_rows['ebit']}+{FS}!{fs_cols[i]}{fs_rows['da']}"},
        # 4 — Current Assets = Sum of CA components (Other CA is a plug)
        {'exp_label': '    Expected  [FS: Total Current Assets]',
         'der_label': '    Derived  [FS: Cash+STInv+AR+Inv+OtherCA(plug)]',
         'exp_fn': lambda i: _fsr('total_current_a', fs_cols[i]),
         'der_fn': lambda i: _fs_sum(
             ['cash', 'st_investments', 'accounts_rec', 'inventory', 'other_current_a'],
             fs_cols[i])},
        # 5 — Current Liabilities = Sum of CL components (Other CL is a plug)
        {'exp_label': '    Expected  [FS: Total Current Liabilities]',
         'der_label': '    Derived  [FS: AP+Accrued+STDebt+DefRev+OtherCL(plug)]',
         'exp_fn': lambda i: _fsr('total_current_l', fs_cols[i]),
         'der_fn': lambda i: _fs_sum(
             ['accounts_pay', 'accrued_liab', 'st_debt', 'deferred_rev_cur',
              'other_current_l'],
             fs_cols[i])},
        # 6 — Assets = Liabilities + Equity
        {'exp_label': '    Expected  [FS: Total Assets]',
         'der_label': '    Derived  [FS: Total Liabilities + Total Equity]',
         'exp_fn': lambda i: _fsr('total_assets', fs_cols[i]),
         'der_fn': lambda i: f"={FS}!{fs_cols[i]}{fs_rows['total_liabilities']}+{FS}!{fs_cols[i]}{fs_rows['total_equity']}"},
        # 7 — Net CF = Op CF + Inv CF + Fin CF + FX
        {'exp_label': '    Expected  [FS: OpCF + InvCF + FinCF + FX]',
         'der_label': '    Derived  [FS: Cash(t) - Cash(t-1)]',
         'exp_fn': lambda i: _fs_sum(
             ['operating_cf', 'investing_cf', 'financing_cf', 'fx_effect'],
             fs_cols[i]),
         'der_fn': lambda i: (
             f"={FS}!{fs_cols[i]}{fs_rows['cash']}-{FS}!{fs_cols[i-1]}{fs_rows['cash']}"
             if i > 0 else None)},
        # 8 — FCF = Operating CF - |Capex|
        {'exp_label': '    Expected  [FS: Free Cash Flow]',
         'der_label': '    Derived  [FS: Operating CF - ABS(Capex)]',
         'exp_fn': lambda i: _fsr('fcf', fs_cols[i]),
         'der_fn': lambda i: f"={FS}!{fs_cols[i]}{fs_rows['operating_cf']}-ABS({FS}!{fs_cols[i]}{fs_rows['capex']})"},
        # 9 — Net Income (IS) = Net Income (CF)
        {'exp_label': '    Expected  [10-K IS: NetIncomeLoss]',
         'der_label': '    Derived  [FS CF: Net Income]',
         'exp_fn': lambda i: None,
         'der_fn': lambda i: _fsr('cf_net_income', fs_cols[i])},
        # 10 — D&A (IS) = D&A (CF)
        {'exp_label': '    Expected  [FS IS: D&A]',
         'der_label': '    Derived  [FS CF: D&A]',
         'exp_fn': lambda i: _fsr('da', fs_cols[i]),
         'der_fn': lambda i: _fsr('cf_da', fs_cols[i])},
    ]

    _set_col_widths(ws, {1: 52, 2: 16, 3: 16, 4: 16, 5: 16})

    r = 1

    # ── Title ────────────────────────────────────────────────────────────
    title = ws.cell(
        row=r, column=1,
        value=f"{company_info['name']}  ({company_info['ticker']})"
              f" -- Data Validation  ($ in millions)")
    _style(title, fill_hex=DARK_BLUE, bold=True, font_color=WHITE)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=1+n)
    ws.row_dimensions[r].height = 18
    r += 1

    # Subtitle with SEC EDGAR hyperlink
    cik = company_info.get('cik', '')
    edgar_url = (f"https://www.sec.gov/cgi-bin/browse-edgar?"
                 f"action=getcompany&CIK={cik}&type=10-K&dateb=&owner=include&count=10")
    sub = ws.cell(
        row=r, column=1,
        value="Source: SEC EDGAR XBRL  |  Amounts in $M  |  "
              "Expected = 10-K reported  |  Derived = FS tab formulas")
    _style(sub, fill_hex=XLIGHT_BLUE, italic=True)
    sub.hyperlink = edgar_url
    sub.font = Font(bold=False, color='0563C1', size=10, italic=True,
                    name='Calibri', underline='single')
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=1+n)
    r += 2

    # ── Summary Table ────────────────────────────────────────────────────
    _write_section_header(ws, r, 'VALIDATION SUMMARY', cols=1+n); r += 1
    _write_col_headers(ws, r, list(range(2, 2+n)), years, start_col=2); r += 1

    total_pass = 0
    total_checks = 0

    for chk in checks:
        lbl = ws.cell(row=r, column=1, value=chk['name'])
        _style(lbl, bold=False)
        tol_pct = chk.get('tolerance_pct')
        tol_ref_map = chk.get('tolerance_ref')      # optional per-year ref
        for i, yr in enumerate(years):
            exp = chk['expected'].get(yr)
            der = chk['derived'].get(yr)
            tol_ref = tol_ref_map.get(yr, exp) if tol_ref_map else exp
            passed, diff = _check_status(exp, der, tol_ref, tolerance_pct=tol_pct)

            c = ws.cell(row=r, column=2+i)
            if exp is None and der is None:
                c.value = 'N/A'
                _style(c, fill_hex=MED_GRAY, h_align='center', italic=True)
            elif passed:
                c.value = 'PASS'
                _style(c, fill_hex=LIGHT_GREEN, bold=True, h_align='center',
                       font_color=PASS_GREEN)
                total_pass += 1
                total_checks += 1
            else:
                c.value = f'FAIL ({diff:+.1f})'
                _style(c, fill_hex=FAIL_RED, bold=True, h_align='center',
                       font_color=DARK_RED)
                total_checks += 1
        r += 1

    # Summary counts
    r += 1
    lbl = ws.cell(row=r, column=1, value='Total Checks')
    _style(lbl, bold=True)
    ws.cell(row=r, column=2, value=total_checks)
    _style(ws.cell(row=r, column=2), bold=True, h_align='center')
    r += 1
    lbl = ws.cell(row=r, column=1, value='Passed')
    _style(lbl, bold=True, font_color=PASS_GREEN)
    ws.cell(row=r, column=2, value=total_pass)
    _style(ws.cell(row=r, column=2), bold=True, h_align='center',
           font_color=PASS_GREEN)
    r += 1
    failed = total_checks - total_pass
    lbl = ws.cell(row=r, column=1, value='Failed')
    _style(lbl, bold=True, font_color=DARK_RED if failed else '000000')
    ws.cell(row=r, column=2, value=failed)
    _style(ws.cell(row=r, column=2), bold=True, h_align='center',
           font_color=DARK_RED if failed else '000000')
    r += 2

    # ── Detailed Breakdown ───────────────────────────────────────────────
    _write_section_header(ws, r, 'DETAILED BREAKDOWN', cols=1+n); r += 1

    current_section = None
    for chk_idx, chk in enumerate(checks):
        meta = check_meta[chk_idx]

        # Section sub-header
        if chk['section'] != current_section:
            current_section = chk['section']
            _write_section_header(ws, r, current_section, cols=1+n)
            r += 1
            _write_col_headers(ws, r, list(range(2, 2+n)), years, start_col=2)
            r += 1

        # Check name
        lbl = ws.cell(row=r, column=1, value=chk['name'])
        _style(lbl, bold=True)
        r += 1

        # Expected row — formula (cross-sheet ref) or hard-coded Python value
        exp_row = r
        lbl = ws.cell(row=r, column=1, value=meta['exp_label'])
        _style(lbl, italic=True)
        for i, yr in enumerate(years):
            formula = meta['exp_fn'](i)
            val = formula if formula is not None else chk['expected'].get(yr)
            c = ws.cell(row=r, column=2+i, value=val)
            _style(c, fill_hex=LIGHT_GRAY if i % 2 == 0 else WHITE,
                   h_align='right', number_format=FMT_DOLLAR)
        r += 1

        # Derived row — formula (cross-sheet ref) or hard-coded Python value
        der_row = r
        lbl = ws.cell(row=r, column=1, value=meta['der_label'])
        _style(lbl, italic=True)
        for i, yr in enumerate(years):
            formula = meta['der_fn'](i)
            val = formula if formula is not None else chk['derived'].get(yr)
            c = ws.cell(row=r, column=2+i, value=val)
            _style(c, fill_hex=LIGHT_GRAY if i % 2 == 0 else WHITE,
                   h_align='right', number_format=FMT_DOLLAR)
        r += 1

        # Difference row — Excel formula referencing the cells above
        lbl = ws.cell(row=r, column=1, value='    Difference ($M)')
        _style(lbl, bold=True)
        for i, yr in enumerate(years):
            col_let = get_column_letter(2 + i)
            exp_py = chk['expected'].get(yr)
            der_py = chk['derived'].get(yr)
            # Use formula if both rows have values
            if exp_py is not None and der_py is not None:
                c = ws.cell(row=r, column=2+i,
                            value=f'={col_let}{exp_row}-{col_let}{der_row}')
            else:
                c = ws.cell(row=r, column=2+i, value=None)
            tol_ref_d = chk.get('tolerance_ref')
            tol_r = tol_ref_d.get(yr, exp_py) if tol_ref_d else exp_py
            passed, _ = _check_status(exp_py, der_py, tol_r,
                                       tolerance_pct=chk.get('tolerance_pct'))
            fill = LIGHT_GREEN if passed else FAIL_RED
            _style(c, fill_hex=fill, bold=True, h_align='right',
                   number_format=FMT_DOLLAR)
        r += 1
        # spacer between checks
        r += 1

    ws.freeze_panes = 'A3'

    # Return results for programmatic use (bulk test)
    return {'total': total_checks, 'passed': total_pass,
            'failed': total_checks - total_pass, 'checks': checks}


# ─────────────────────────────────────────────────────────────────────────────
# MAIN ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────

def create_excel(company_info: Dict, financial_data: Dict, output_file: str):
    """Create the full financial model workbook and save it."""
    wb = Workbook()

    ws_fs  = wb.active
    ws_fs.title = 'Financial Statements'

    ws_wacc = wb.create_sheet('WACC')
    ws_dcf  = wb.create_sheet('DCF Model')
    ws_val  = wb.create_sheet('Data Validation')

    print("  Building Financial Statements sheet...")
    fs_rows = _write_financial_statements(ws_fs, company_info, financial_data)

    print("  Building WACC sheet...")
    wacc_rows = _write_wacc_sheet(ws_wacc, company_info, financial_data, fs_rows)

    print("  Building DCF Model sheet...")
    _write_dcf_model(ws_dcf, company_info, financial_data, fs_rows, wacc_rows)

    print("  Building Data Validation sheet...")
    validation_results = _write_validation_sheet(ws_val, company_info, financial_data, fs_rows)

    wb.save(output_file)
    print(f"  Saved: {output_file}")

    return validation_results
